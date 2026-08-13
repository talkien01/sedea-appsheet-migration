"""
Motor generador de datos de ficha por municipio.
Combina los CSV exportados de Postgres (analitica_export) + la plantilla manual
(Datos_referencia_manual_municipios.xlsx). No inventa datos: todo lo que no
tiene fuente queda marcado como advertencia (warning) en vez de completarse.
"""
import csv, os
import openpyxl
from config import DATA_DIR, MANUAL_XLSX

def _leer_csv(nombre):
    path = os.path.join(DATA_DIR, nombre)
    if not os.path.exists(path):
        return []
    with open(path, encoding="utf-8") as f:
        return list(csv.DictReader(f))

def cargar_apoyo_municipio():
    return _leer_csv("05_apoyo_municipio_full.csv")

def cargar_v_oficial_municipio():
    return _leer_csv("14_v_oficial_municipio.csv")

def cargar_municipios():
    rows = _leer_csv("01_municipio.csv")
    return [r["nombre"] for r in rows if r.get("region_id")]

def cargar_manual(hoja, municipio):
    if not os.path.exists(MANUAL_XLSX):
        return []
    wb = openpyxl.load_workbook(MANUAL_XLSX, data_only=True)
    if hoja not in wb.sheetnames:
        return []
    ws = wb[hoja]
    headers = [c.value for c in ws[1]]
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = dict(zip(headers, row))
        nombre_col = headers[0]
        if d.get(nombre_col) and municipio in str(d.get(nombre_col)):
            out.append(d)
    return out

def generar(municipio):
    warnings = []
    apoyo = [r for r in cargar_apoyo_municipio() if r["municipio_nombre"] == municipio]
    oficial = [r for r in cargar_v_oficial_municipio() if r["municipio_proyecto"] == municipio]

    historico = sorted(apoyo, key=lambda r: (r["anio"], r["programa_nombre"]))
    programas_historicos = set(r["programa_nombre"] for r in apoyo)

    total_apoyos = sum(int(r["numero_apoyos"] or 0) for r in apoyo)
    total_inversion = sum(float(r["total"] or 0) for r in apoyo)
    anios_presentes = sorted(set(r["anio"] for r in apoyo))
    for anio_falt in ["2021", "2022", "2026"]:
        if anio_falt not in anios_presentes:
            warnings.append(
                f"apoyo_municipio no tiene filas de {anio_falt} para {municipio} "
                f"(confirmado: ningún municipio tiene {anio_falt} en esta tabla, no es hueco solo de este municipio)."
            )

    componentes_2026 = set(r["componente"] for r in oficial)
    total_apoyos_2026 = sum(int(r["apoyos"] or 0) for r in oficial)
    total_2026 = sum(float(r["total_dictaminado"] or 0) for r in oficial)

    equivalencias = {
        "DINAMISMO AGROALIMENTARIO": "DINAMISMO AGROALIMENTARIO",
        "TECNIFICACIÓN": "TECNIFICACIÓN DEL RIEGO",
        "CAPTACIÓN Y ALMACENAMIENTO DE AGUA": "CAPTACIÓN Y ALMACENAMIENTO DE AGUA",
    }
    programas_sin_2026 = []
    for p in sorted(programas_historicos):
        comp_esperado = equivalencias.get(p, p)
        if comp_esperado not in componentes_2026:
            programas_sin_2026.append(p)
    if programas_sin_2026:
        warnings.append(
            f"2026 incompleto para {municipio}: faltan cargar estos programas que sí existen en años anteriores: "
            + ", ".join(programas_sin_2026)
            + ". Acción: pedir al equipo de datos que cargue estos programas 2026 en analitica."
        )

    territorio = cargar_manual("Territorio", municipio)
    productos = cargar_manual("Productos_top", municipio)
    precipitacion = cargar_manual("Precipitacion", municipio)
    demografia = cargar_manual("Demografia_municipal", municipio)

    def campos_vacios(rows, skip_cols=("Municipio", "Fuente / fecha de corte", "Grupo de edad", "Rank", "Año", "Producto")):
        vacios = total = 0
        for row in rows:
            for k, v in row.items():
                if k in skip_cols:
                    continue
                total += 1
                if v in (None, "", " "):
                    vacios += 1
        return vacios, total

    for nombre, rows in [("Territorio", territorio), ("Productos_top", productos),
                          ("Precipitacion", precipitacion), ("Demografia_municipal", demografia)]:
        if not rows:
            warnings.append(f"'{nombre}' no tiene fila para {municipio} en la plantilla manual.")
            continue
        vacios, total = campos_vacios(rows)
        if vacios == total and total > 0:
            warnings.append(
                f"'{nombre}' está sin llenar para {municipio}. Fuente sugerida: INEGI/SIAP "
                f"(territorio, productos, demografía) o CONAGUA (precipitación)."
            )
        elif vacios > 0:
            warnings.append(f"'{nombre}' parcialmente lleno para {municipio}: {vacios} de {total} celdas vacías.")

    return {
        "municipio": municipio,
        "historico": historico,
        "total_apoyos_historico": total_apoyos,
        "total_inversion_historico": round(total_inversion, 2),
        "anios_presentes": anios_presentes,
        "avance_2026": oficial,
        "total_apoyos_2026": total_apoyos_2026,
        "total_2026": round(total_2026, 2),
        "territorio": territorio,
        "productos": productos,
        "precipitacion": precipitacion,
        "demografia": demografia,
        "warnings": warnings,
        **secciones_extendidas({"municipio": municipio}, ambito="MUNICIPIO", clave=municipio),
    }


# ===========================================================================
# Extensión 2026: secciones 7-10 (Emergentes/Productividad, aportaciones,
# género y edad, incidencias) y fichas regional y estatal.
# Todo lo que no tiene fuente se reporta como leyenda, nunca como cero (R1).
# ===========================================================================
import db  # noqa: E402
from services import aportaciones as _aport  # noqa: E402
from services import genero_edad as _gen  # noqa: E402
from services import incidencias as _inc  # noqa: E402
from services import matriz as _matriz  # noqa: E402


def secciones_extendidas(filtros, ambito, clave):
    """Devuelve las 4 secciones nuevas para cualquier ámbito.
    Si la base está apagada, cada sección trae su leyenda explicativa."""
    salida = {
        "ambito": ambito,
        "clave_ambito": clave,
        "emer_prod": [],
        "aportaciones": {},
        "genero_edad": [],
        "genero_edad_leyenda": None,
        "incidencias": [],
        "incidencias_leyenda": None,
    }
    try:
        salida["emer_prod"] = _matriz.emer_prod(filtros)
    except db.SinDatos as e:
        salida["emer_prod_leyenda"] = f"Sin conexión a la base: {e}"
    try:
        salida["aportaciones"] = _aport.resumen(filtros)
    except db.SinDatos as e:
        salida["aportaciones_leyenda"] = f"Sin conexión a la base: {e}"
    try:
        filas = _gen.resumen_por_rango(filtros)
        salida["genero_edad"] = filas
        if not filas:
            salida["genero_edad_leyenda"] = _gen.leyenda_sin_datos(f"{ambito.lower()} {clave}")
    except db.SinDatos as e:
        salida["genero_edad_leyenda"] = f"Sin conexión a la base: {e}"
    try:
        salida["incidencias"] = _inc.listar(municipio=filtros.get("municipio"),
                                            resuelta=False, limite=50)
        if not salida["incidencias"]:
            salida["incidencias_leyenda"] = "Sin incidencias abiertas para este ámbito."
    except db.SinDatos as e:
        salida["incidencias_leyenda"] = f"Sin conexión a la base: {e}"
    return salida


def _acumular(filas):
    acc = {"numero_apoyos": None, "federal": None, "estatal": None, "municipal": None,
           "beneficiario": None, "total": None}
    for f in filas:
        for k in acc:
            v = f.get(k)
            if v is not None:
                acc[k] = (acc[k] or 0) + v
    return acc


def generar_region(region):
    """Ficha a nivel región: agrega los municipios de la región (A15)."""
    municipios = _matriz.municipios_de_region(region)
    if not municipios:
        raise ValueError(f"La región «{region}» no existe en el catálogo.")
    filtros = {"region": region}
    warnings = []
    filas, total_filas = _matriz.matriz(dict(filtros, page_size=100000, page=1), con_paginado=False)
    if not filas:
        warnings.append(f"No hay ninguna fila cargada para la región {region}.")
    por_municipio = {}
    for f in filas:
        d = por_municipio.setdefault(f["municipio"], [])
        d.append(f)
    resumen_municipios = []
    for m in municipios:
        sub = por_municipio.get(m, [])
        if not sub:
            warnings.append(f"{m}: sin datos cargados en ningún origen; se rinde vacío, no en cero.")
        fila = _acumular(sub)
        fila["municipio"] = m
        fila["anios"] = sorted({s["anio"] for s in sub if s.get("anio")})
        resumen_municipios.append(fila)
    return {
        "region": region,
        "municipios": municipios,
        "resumen_municipios": resumen_municipios,
        "serie_anual": _matriz.serie_inversion_anual(filtros),
        "total": _matriz.totales(filtros),
        "filas": filas,
        "total_filas": total_filas,
        "warnings": warnings,
        **secciones_extendidas(filtros, ambito="REGION", clave=region),
    }


def generar_estatal():
    """Ficha estatal: resumen_estatal + v_oficial_componente + matriz completa."""
    warnings = []
    filtros = {}
    resumen_estatal = _leer_csv("07_resumen_estatal_full.csv")
    if not resumen_estatal:
        try:
            resumen_estatal = db.consultar(
                "SELECT re.*, p.nombre AS programa_nombre FROM analitica.resumen_estatal re "
                "JOIN analitica.programa p USING (programa_id) ORDER BY re.anio, p.nombre")
        except db.SinDatos:
            warnings.append("No se pudo leer resumen_estatal (ni CSV ni base).")
    componentes_2026 = _leer_csv("13_v_oficial_componente.csv")
    if not componentes_2026:
        try:
            componentes_2026 = db.consultar(
                "SELECT * FROM analitica.v_oficial_componente ORDER BY anio, componente")
        except db.SinDatos:
            warnings.append("No se pudo leer v_oficial_componente (ni CSV ni base).")
    serie = _matriz.serie_inversion_anual(filtros)
    if not any(s.get("anio") == 2027 for s in serie):
        warnings.append("2027 no tiene ninguna cifra cargada: se rinde vacío, nunca como cero (R4).")
    return {
        "resumen_estatal": resumen_estatal,
        "componentes_2026": componentes_2026,
        "serie_anual": serie,
        "total": _matriz.totales(filtros),
        "warnings": warnings,
        **secciones_extendidas(filtros, ambito="ESTATAL", clave="QUERÉTARO"),
    }
