"""Carga analitica.beneficiario_curp desde el padrón por folio.

    python -m ingesta.cargar_curp --archivo "Base Cadereyta Programas sedea.xlsx" --dry-run
    python -m ingesta.cargar_curp --archivo "Base Cadereyta Programas sedea.xlsx"
    python -m ingesta.cargar_curp --desde-oficial            # 2026 estatal desde oficial.solicitud

El género, la fecha de nacimiento y el rango de edad se derivan **solo** de la
CURP (R6). Si la CURP no es válida quedan en NULL y se levanta incidencia
CURP_INVALIDA. El nombre de la persona nunca interviene.
"""
import argparse
import sys

import openpyxl

import config
import db
from ingesta import comun
from services import curp as svc_curp
from services import incidencias

MAPEO = "curp_base_cadereyta.json"

_SQL_INSERT = """
INSERT INTO analitica.beneficiario_curp
  (curp_hash, curp, curp_valida, motivo_invalidez, genero, fecha_nacimiento,
   edad_anios, rango_edad, entidad_nacimiento, anio, programa_id, municipio_id,
   municipio_usado, fuente_municipio, folio, monto_total,
   fuente_archivo, fuente_hoja, fila_origen, fecha_corte)
VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
ON CONFLICT (curp_hash, anio, coalesce(programa_id,-1), coalesce(folio,''))
DO UPDATE SET
  curp = EXCLUDED.curp,
  curp_valida = EXCLUDED.curp_valida,
  motivo_invalidez = EXCLUDED.motivo_invalidez,
  genero = EXCLUDED.genero,
  fecha_nacimiento = EXCLUDED.fecha_nacimiento,
  edad_anios = EXCLUDED.edad_anios,
  rango_edad = EXCLUDED.rango_edad,
  entidad_nacimiento = EXCLUDED.entidad_nacimiento,
  municipio_id = EXCLUDED.municipio_id,
  municipio_usado = EXCLUDED.municipio_usado,
  fuente_municipio = EXCLUDED.fuente_municipio,
  monto_total = EXCLUDED.monto_total,
  fuente_archivo = EXCLUDED.fuente_archivo,
  fuente_hoja = EXCLUDED.fuente_hoja,
  fila_origen = EXCLUDED.fila_origen,
  fecha_corte = EXCLUDED.fecha_corte,
  cargado_en = now()
"""


def _contar():
    return int(db.escalar("SELECT count(*)::int FROM analitica.beneficiario_curp", default=0) or 0)


def _volcar(registros, contadores, dry_run):
    """Escribe el lote completo en una sola conexión (idempotente por UPSERT).
    insertadas = cuánto creció la tabla; el resto son actualizaciones."""
    if dry_run:
        contadores.omitidas += len(registros)
        return
    antes = _contar()
    db.ejecutar_muchos(_SQL_INSERT, registros)
    despues = _contar()
    contadores.insertadas += despues - antes
    contadores.actualizadas += len(registros) - (despues - antes)


def cargar_xlsx(archivo, dry_run=False, anio=None):
    mapeo = comun.cargar_mapeo(MAPEO)
    ruta = comun.ruta_archivo(archivo)
    nombre_archivo = archivo
    wb = openpyxl.load_workbook(ruta, data_only=True)
    hoja = mapeo["hoja"] if mapeo["hoja"] in wb.sheetnames else wb.sheetnames[0]
    ws = wb[hoja]

    filas = list(ws.iter_rows(values_only=True))
    fila_enc = mapeo.get("fila_encabezado", 1)
    encabezados = filas[fila_enc - 1]
    indice = comun.indexar_encabezados(encabezados)
    resuelto, faltantes = comun.resolver_mapeo(indice, mapeo["columnas"])

    print(f"Archivo : {nombre_archivo}")
    print(f"Hoja    : {hoja}")
    print("Encabezados detectados:", [h for h in encabezados if h])
    print("Columnas mapeadas     :",
          {k: (encabezados[v] if v is not None else None) for k, v in resuelto.items()})
    print("Encabezados NO mapeados:", comun.no_mapeadas(indice, resuelto))
    if resuelto.get("curp") is None:
        print(f"ERROR: la hoja «{hoja}» de este archivo no tiene columna CURP; no se puede "
              f"derivar género ni edad sin inventar (R1/R6). Usa el padrón por folio.",
              file=sys.stderr)
        return 3
    if faltantes:
        print("Campos sin columna en el origen (quedarán NULL):", faltantes)

    por_nombre, alias, pseudo = comun.catalogo_municipios(db)
    prog_nombre, prog_alias = comun.catalogo_programas(db)
    corte = config.fecha_corte_date()
    filtro = mapeo.get("filtro_estatus") or {}
    idx_estatus = resuelto.get("estatus")
    c = comun.Contadores()
    invalidas = []
    registros = []

    for n_fila, fila in enumerate(filas[fila_enc:], start=fila_enc + 1):
        if all(v in (None, "") for v in fila):
            continue
        c.leidas += 1
        if idx_estatus is not None and filtro.get("valores_aceptados"):
            estatus = comun.normalizar(fila[idx_estatus])
            if estatus not in [comun.normalizar(v) for v in filtro["valores_aceptados"]]:
                c.omitidas += 1
                continue

        crudo = fila[resuelto["curp"]]
        an = svc_curp.analizar(crudo, corte)
        literal_muni = fila[resuelto["municipio"]] if resuelto.get("municipio") is not None else None
        muni_id, fuente_muni, _conf = comun.resolver_municipio(literal_muni, por_nombre, alias, pseudo)
        prog_id = (comun.resolver_programa(fila[resuelto["programa"]], prog_nombre, prog_alias)
                   if resuelto.get("programa") is not None else None)
        anio_fila = anio or (comun.a_entero(fila[resuelto["anio"]])
                             if resuelto.get("anio") is not None else None) \
            or mapeo.get("anio_por_defecto")
        folio = (str(fila[resuelto["folio"]]).strip()
                 if resuelto.get("folio") is not None and fila[resuelto["folio"]] else None)
        monto = (comun.a_numero(fila[resuelto["monto_total"]])
                 if resuelto.get("monto_total") is not None else None)

        registro = (
            an["curp_hash"], an["curp"], an["curp_valida"], an["motivo_invalidez"],
            an["genero"], an["fecha_nacimiento"], an["edad_anios"], an["rango_edad"],
            an["entidad_nacimiento"], anio_fila, prog_id, muni_id,
            str(literal_muni) if literal_muni else "(sin municipio en el origen)",
            fuente_muni, folio, monto, nombre_archivo, hoja, n_fila, corte,
        )
        registros.append(registro)

        if not an["curp_valida"]:
            invalidas.append((n_fila, an["motivo_invalidez"], folio, muni_id, anio_fila))
        if muni_id is None and not dry_run:
            incidencias.registrar(
                "MUNICIPIO_NO_RESUELTO", "ADVERTENCIA", "beneficiario_curp",
                descripcion=f"Fila {n_fila}: municipio «{literal_muni}» no está en el catálogo ni en los alias.",
                accion_sugerida="Agregar el alias en analitica.municipio_alias y recargar.",
                anio=anio_fila, valor_origen=str(literal_muni),
                fuente_archivo=nombre_archivo, fuente_hoja=hoja, fila_origen=n_fila)
            c.incidencias += 1

    _volcar(registros, c, dry_run)

    if invalidas and not dry_run:
        for n_fila, motivo, folio, muni_id, anio_fila in invalidas:
            incidencias.registrar(
                "CURP_INVALIDA", "ADVERTENCIA", "beneficiario_curp",
                descripcion=(f"Fila {n_fila} (folio {folio}): la CURP no pasó la validación "
                             f"({motivo}); género, fecha de nacimiento y rango de edad quedan vacíos."),
                accion_sugerida=("Corregir la CURP en el padrón de origen y recargar. Está prohibido "
                                 "inferir sexo o edad por otros medios (R6)."),
                anio=anio_fila, municipio_id=muni_id, valor_origen=motivo,
                fuente_archivo=nombre_archivo, fuente_hoja=hoja, fila_origen=n_fila)
            c.incidencias += 1

    c.imprimir(dry_run)
    return 0


def cargar_desde_oficial(dry_run=False, anio=2026):
    """Padrón 2026 completo que ya vive en la base (oficial.solicitud), con su
    archivo fuente real propagado a fuente_archivo."""
    filas = db.consultar(
        "SELECT id, folio, anio, componente, curp, municipio_proyecto, id_municipio, "
        "       monto_total_dictaminado, fuente_archivo "
        "FROM oficial.solicitud WHERE anio = %s AND curp IS NOT NULL", (anio,))
    print(f"Origen  : oficial.solicitud (anio={anio})")
    print(f"Filas con CURP: {len(filas)}")
    por_nombre, alias, pseudo = comun.catalogo_municipios(db)
    prog_nombre, prog_alias = comun.catalogo_programas(db)
    corte = config.fecha_corte_date()
    c = comun.Contadores()
    registros = []
    for f in filas:
        c.leidas += 1
        an = svc_curp.analizar(f["curp"], corte)
        muni_id, fuente_muni, _ = comun.resolver_municipio(
            f["municipio_proyecto"], por_nombre, alias, pseudo)
        prog_id = comun.resolver_programa(f["componente"], prog_nombre, prog_alias)
        registro = (
            an["curp_hash"], an["curp"], an["curp_valida"], an["motivo_invalidez"],
            an["genero"], an["fecha_nacimiento"], an["edad_anios"], an["rango_edad"],
            an["entidad_nacimiento"], int(f["anio"]), prog_id, muni_id,
            f["municipio_proyecto"] or "(sin municipio en el origen)", fuente_muni,
            f["folio"], comun.a_numero(f["monto_total_dictaminado"]),
            f["fuente_archivo"] or "oficial.solicitud", "oficial.solicitud",
            int(f["id"]), corte,
        )
        registros.append(registro)
        if not an["curp_valida"] and not dry_run:
            incidencias.registrar(
                "CURP_INVALIDA", "ADVERTENCIA", "beneficiario_curp",
                descripcion=(f"oficial.solicitud id={f['id']} (folio {f['folio']}): CURP inválida "
                             f"({an['motivo_invalidez']}); derivados vacíos."),
                accion_sugerida="Corregir la CURP en el padrón oficial y recargar (R6).",
                anio=int(f["anio"]), municipio_id=muni_id, valor_origen=an["motivo_invalidez"],
                fuente_archivo=f["fuente_archivo"], fuente_hoja="oficial.solicitud",
                fila_origen=int(f["id"]))
            c.incidencias += 1
    _volcar(registros, c, dry_run)
    c.imprimir(dry_run)
    return 0


def main(argv=None):
    ap = argparse.ArgumentParser(description="Carga beneficiario_curp desde un padrón por folio.")
    ap.add_argument("--archivo", help="ruta del .xlsx con CURP por beneficiario")
    ap.add_argument("--desde-oficial", action="store_true",
                    help="carga desde oficial.solicitud (padrón 2026 ya en la base)")
    ap.add_argument("--anio", type=int, help="ejercicio a asignar/filtrar")
    ap.add_argument("--dry-run", action="store_true", help="no escribe nada en la base")
    args = ap.parse_args(argv)

    if not args.archivo and not args.desde_oficial:
        ap.print_help()
        return 0
    try:
        if args.desde_oficial:
            return cargar_desde_oficial(dry_run=args.dry_run, anio=args.anio or 2026)
        return cargar_xlsx(args.archivo, dry_run=args.dry_run, anio=args.anio)
    except FileNotFoundError as e:
        print(f"ERROR: {e}", file=sys.stderr)
        return 4
    except db.SinDatos as e:
        print(f"ERROR: no hay conexión a la base ({e}).", file=sys.stderr)
        return 2


if __name__ == "__main__":
    sys.exit(main())
