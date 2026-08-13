"""Contrasta la distribución declarada por género/edad contra lo derivado de CURP.

    python -m ingesta.cargar_distribucion --archivo "Regional_Cadereyta_Distribucion_Cadereyta_Colon_Ezequiel.xlsx" --dry-run
    python -m ingesta.cargar_distribucion --archivo "…" --anio 2026

Por qué no se carga como dato: el archivo agrupa la edad en `Adultos (30-59)`,
que **no** es uno de los 5 rangos normalizados; repartirlo entre `30-44` y
`45-59` sería inventar (R1). La fuente primaria de género y edad es
`analitica.beneficiario_curp` (R6). Este script compara ambos y levanta
incidencia cuando difieren.
"""
import argparse
import sys

import openpyxl

import db
from ingesta import comun
from services import incidencias

MAPEO = "distribucion_genero_edad.json"


def _bloques(ws, municipios_norm):
    """Devuelve [{municipio, fila, hombres, mujeres, total}] leyendo los bloques."""
    filas = list(ws.iter_rows(values_only=True))
    bloques, actual = [], None
    for n, fila in enumerate(filas, start=1):
        etiqueta = fila[0]
        if etiqueta and comun.normalizar(etiqueta) in municipios_norm:
            actual = {"municipio": str(etiqueta).strip(), "fila": n,
                      "hombres": None, "mujeres": None, "total": None}
            bloques.append(actual)
            continue
        if actual is None or not etiqueta:
            continue
        clave = comun.normalizar(etiqueta)
        if clave == "HOMBRES" and actual["hombres"] is None:
            actual["hombres"] = comun.a_entero(fila[1])
        elif clave == "MUJERES" and actual["mujeres"] is None:
            actual["mujeres"] = comun.a_entero(fila[1])
        elif clave == "TOTAL" and actual["total"] is None:
            actual["total"] = comun.a_entero(fila[1])
    return bloques


def main(argv=None):
    ap = argparse.ArgumentParser(description="Contraste de distribución género/edad contra CURP.")
    ap.add_argument("--archivo", required=True)
    ap.add_argument("--anio", type=int, default=2026)
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args(argv)

    try:
        mapeo = comun.cargar_mapeo(MAPEO)
        ruta = comun.ruta_archivo(args.archivo)
        wb = openpyxl.load_workbook(ruta, data_only=True)
        hoja = mapeo["hoja"] if mapeo["hoja"] in wb.sheetnames else wb.sheetnames[0]
        ws = wb[hoja]
        print(f"Archivo : {args.archivo}")
        print(f"Hoja    : {hoja}")
        print("Grupos de edad del archivo:", list(mapeo["mapeo_rango_edad"].keys()))
        print("NOTA: 'Adultos (30-59)' no es un rango normalizado; este archivo se usa como "
              "contraste, no como fuente (R1).")

        por_nombre, alias, pseudo = comun.catalogo_municipios(db)
        bloques = _bloques(ws, set(por_nombre) | set(alias))
        c = comun.Contadores()
        for b in bloques:
            c.leidas += 1
            muni_id, _f, _c = comun.resolver_municipio(b["municipio"], por_nombre, alias, pseudo)
            if muni_id is None:
                c.omitidas += 1
                continue
            filas = db.consultar(
                "SELECT genero, count(DISTINCT curp_hash)::int AS n "
                "FROM analitica.beneficiario_curp WHERE anio=%s AND municipio_id=%s AND curp_valida "
                "GROUP BY genero", (args.anio, muni_id))
            derivado = {f["genero"]: int(f["n"]) for f in filas}
            h_curp, m_curp = derivado.get("H"), derivado.get("M")
            print(f"  {b['municipio']}: archivo H={b['hombres']} M={b['mujeres']} | "
                  f"CURP H={h_curp} M={m_curp}")
            if (h_curp, m_curp) != (b["hombres"], b["mujeres"]):
                if not args.dry_run:
                    incidencias.registrar(
                        "DATO_MANUAL_FALTANTE", "INFO", "distribucion_xlsx",
                        descripcion=(f"{b['municipio']} {args.anio}: el archivo de distribución "
                                     f"declara H={b['hombres']}/M={b['mujeres']} y lo derivado de "
                                     f"CURP da H={h_curp}/M={m_curp}."),
                        accion_sugerida=("Conciliar el universo de folios de ambos documentos; la "
                                         "cifra publicable es la derivada de CURP (R6)."),
                        anio=args.anio, municipio_id=muni_id,
                        valor_origen=f"H={b['hombres']};M={b['mujeres']}",
                        fuente_archivo=args.archivo, fuente_hoja=hoja, fila_origen=b["fila"])
                    c.incidencias += 1
            else:
                c.actualizadas += 1
        c.imprimir(args.dry_run)
    except FileNotFoundError as e:
        print(f"ERROR: {e}", file=sys.stderr)
        return 4
    except db.SinDatos as e:
        print(f"ERROR: sin base de datos ({e}).", file=sys.stderr)
        return 2
    return 0


if __name__ == "__main__":
    sys.exit(main())
