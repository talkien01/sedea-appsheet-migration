"""Carga el padrón folio por folio del Programa Institucional Emergente por
Sequía para Productores del Campo (ejercicio 2023, maíz para consumo humano).

    python -m ingesta.cargar_emergentes_sequia --dry-run
    python -m ingesta.cargar_emergentes_sequia
    python -m ingesta.cargar_emergentes_sequia --archivo SDAproyectos23cadSequia.xlsx

Fuente primaria elegida: las 4 hojas consolidadas por región del Drive de SEDEA
(`SDAproyectos23{cad,jal,qro,sjr}Sequía`, pestaña `23_Emer_Sequía_<REG>`), exportadas
a .xlsx. Ver docs/INVENTARIO_FUENTES.md para el porqué frente a los ~20 archivos
divididos por municipio y lote de captura.

Escribe en dos niveles, sin duplicar:
  * `analitica.beneficiario_curp` -> una fila por folio (grano individual, R2).
  * `analitica.apoyo_municipio`   -> el agregado municipio x programa x año, con
    `granularidad='FOLIO'` porque se derivó del padrón individual.

Nada se rellena (R1): las columnas de aportación que el padrón trae vacías se
escriben con el valor del dictamen oficial 2023-00007, que las autoriza en $0.00,
y ese origen queda escrito en `observaciones`.
"""
import argparse
import collections
import glob
import os
import sys

import openpyxl

import config
import db
from ingesta import comun
from services import curp as svc_curp
from services import incidencias

MAPEO = "sequia_2023_sipros.json"

FUENTE_DICTAMEN = ("Oficio de autorización 2023GEQ00040, proyecto 2023-00007, "
                   "Secretaría de Finanzas del Estado de Querétaro, 24 de enero de 2023")
MONTO_AUTORIZADO = 45570275.00

_SQL_CURP = """
INSERT INTO analitica.beneficiario_curp
  (curp_hash, curp, curp_valida, motivo_invalidez, genero, fecha_nacimiento,
   edad_anios, rango_edad, entidad_nacimiento, anio, programa_id, municipio_id,
   municipio_usado, fuente_municipio, folio, monto_total,
   fuente_archivo, fuente_hoja, fila_origen, fecha_corte)
VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
ON CONFLICT (curp_hash, anio, coalesce(programa_id,-1), coalesce(folio,''))
DO UPDATE SET
  curp = EXCLUDED.curp, curp_valida = EXCLUDED.curp_valida,
  motivo_invalidez = EXCLUDED.motivo_invalidez, genero = EXCLUDED.genero,
  fecha_nacimiento = EXCLUDED.fecha_nacimiento, edad_anios = EXCLUDED.edad_anios,
  rango_edad = EXCLUDED.rango_edad, entidad_nacimiento = EXCLUDED.entidad_nacimiento,
  municipio_id = EXCLUDED.municipio_id, municipio_usado = EXCLUDED.municipio_usado,
  fuente_municipio = EXCLUDED.fuente_municipio, monto_total = EXCLUDED.monto_total,
  fuente_archivo = EXCLUDED.fuente_archivo, fuente_hoja = EXCLUDED.fuente_hoja,
  fila_origen = EXCLUDED.fila_origen, fecha_corte = EXCLUDED.fecha_corte,
  cargado_en = now()
"""

_SQL_APOYO = """
INSERT INTO analitica.apoyo_municipio
  (anio, programa_id, municipio_id, numero_apoyos, apoyo_federal, apoyo_estatal,
   apoyo_municipal, aportacion_productor, total, fuente_archivo, fuente_hoja,
   municipio_usado, fuente_municipio, confianza_municipio, granularidad, observaciones)
VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
ON CONFLICT (anio, programa_id, municipio_id) DO UPDATE SET
  numero_apoyos = EXCLUDED.numero_apoyos,
  apoyo_federal = EXCLUDED.apoyo_federal,
  apoyo_estatal = EXCLUDED.apoyo_estatal,
  apoyo_municipal = EXCLUDED.apoyo_municipal,
  aportacion_productor = EXCLUDED.aportacion_productor,
  total = EXCLUDED.total,
  fuente_archivo = EXCLUDED.fuente_archivo,
  fuente_hoja = EXCLUDED.fuente_hoja,
  municipio_usado = EXCLUDED.municipio_usado,
  fuente_municipio = EXCLUDED.fuente_municipio,
  confianza_municipio = EXCLUDED.confianza_municipio,
  granularidad = EXCLUDED.granularidad,
  observaciones = EXCLUDED.observaciones,
  cargado_en = now()
"""


def _archivos(mapeo, archivo=None):
    if archivo:
        return [comun.ruta_archivo(archivo)]
    rutas = []
    for nombre in mapeo["fuente_primaria"]:
        try:
            rutas.append(comun.ruta_archivo(nombre))
        except FileNotFoundError:
            print(f"AVISO: no se encontró {nombre}; se omite.", file=sys.stderr)
    if not rutas:
        raise FileNotFoundError(
            "Ninguna de las hojas consolidadas por región de Sequía 2023 está en "
            f"{config.BASE_DIR}. Esperadas: " + ", ".join(mapeo["fuente_primaria"]))
    return rutas


def _hoja(wb, mapeo):
    patron = comun.normalizar(mapeo.get("hoja_patron", ""))
    for nombre in wb.sheetnames:
        if patron and patron in comun.normalizar(nombre):
            return nombre
    return wb.sheetnames[0]


def leer(archivos, mapeo, verbose=True):
    """Devuelve (registros_folio, agregado, contadores, incidencias_pendientes)."""
    por_nombre, alias, pseudo = comun.catalogo_municipios(db)
    prog_nombre, prog_alias = comun.catalogo_programas(db)
    corte = config.fecha_corte_date()
    aceptados = [comun.normalizar(v)
                 for v in (mapeo.get("filtro_estatus") or {}).get("valores_aceptados", [])]

    c = comun.Contadores()
    registros = []
    agregado = collections.defaultdict(
        lambda: {"apoyos": 0, "federal": 0.0, "estatal": 0.0, "municipal": 0.0,
                 "beneficiario": 0.0, "total": 0.0, "municipio_usado": None,
                 "fuente_municipio": None, "confianza": None,
                 "fuente_archivo": None, "fuente_hoja": None})
    pendientes = []           # incidencias a registrar solo si no es dry-run
    vistos = {}               # (folio, municipio_usado) -> primera fila, para FOLIO_DUPLICADO

    for ruta in archivos:
        nombre_archivo = os.path.basename(ruta)
        wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
        hoja = _hoja(wb, mapeo)
        ws = wb[hoja]
        it = ws.iter_rows(values_only=True)
        encabezados = list(next(it))
        indice = comun.indexar_encabezados(encabezados)
        resuelto, faltantes = comun.resolver_mapeo(indice, mapeo["columnas"])

        if verbose:
            print(f"\nArchivo : {nombre_archivo}")
            print(f"Hoja    : {hoja}")
            print(f"Encabezados detectados: {len(indice)}")
            print("Columnas mapeadas     :",
                  {k: (encabezados[v] if v is not None else None) for k, v in resuelto.items()})
            if faltantes:
                print("Campos sin columna en el origen (quedarán NULL):", faltantes)
        if resuelto.get("curp") is None or resuelto.get("folio") is None:
            raise SystemExit(
                f"ERROR: la hoja «{hoja}» de {nombre_archivo} no tiene CURP y/o FOLIO; "
                "no es el padrón folio por folio (R1/R6).")

        def val(fila, campo):
            i = resuelto.get(campo)
            return fila[i] if i is not None and i < len(fila) else None

        for n_fila, fila in enumerate(it, start=mapeo.get("fila_encabezado", 1) + 1):
            if fila is None or all(v in (None, "") for v in fila):
                continue
            folio_crudo = val(fila, "folio")
            if not folio_crudo:
                continue
            c.leidas += 1

            if aceptados and comun.normalizar(val(fila, "dictamen")) not in aceptados:
                c.omitidas += 1
                continue

            folio = str(folio_crudo).strip()
            literal_muni = val(fila, "municipio")
            muni_id, fuente_muni, conf = comun.resolver_municipio(
                literal_muni, por_nombre, alias, pseudo)
            muni_usado = str(literal_muni).strip() if literal_muni else "(sin municipio en el origen)"

            prog_literal = val(fila, "programa") or mapeo["programa_por_defecto"]
            prog_id = comun.resolver_programa(prog_literal, prog_nombre, prog_alias)
            anio = comun.a_entero(val(fila, "anio")) or mapeo["anio_por_defecto"]

            estatal = comun.a_numero(val(fila, "apoyo_estatal"))
            total = comun.a_numero(val(fila, "monto_total"))
            # El dictamen 2023-00007 autoriza federal/municipal/otros en $0.00: el
            # padrón los deja vacíos porque no hubo esa concurrencia. No es imputación.
            federal = comun.a_numero(val(fila, "apoyo_federal")) or 0.0
            municipal = comun.a_numero(val(fila, "apoyo_municipal")) or 0.0
            productor = comun.a_numero(val(fila, "aportacion_productor")) or 0.0

            an = svc_curp.analizar(val(fila, "curp"), corte)
            registros.append((
                an["curp_hash"], an["curp"], an["curp_valida"], an["motivo_invalidez"],
                an["genero"], an["fecha_nacimiento"], an["edad_anios"], an["rango_edad"],
                an["entidad_nacimiento"], anio, prog_id, muni_id,
                muni_usado, fuente_muni, folio, total,
                nombre_archivo, hoja, n_fila, corte,
            ))

            clave = (folio, comun.normalizar(muni_usado))
            if clave in vistos:
                pendientes.append((
                    "FOLIO_DUPLICADO", "ADVERTENCIA", "beneficiario_curp",
                    f"El folio «{folio}» del municipio «{muni_usado}» aparece más de una vez "
                    f"(filas {vistos[clave]} y {n_fila} de {nombre_archivo}).",
                    "Verificar con la Dirección Regional cuál de los dos registros es el válido; "
                    "no se deduplica automáticamente para no perder un apoyo real.",
                    anio, muni_id, prog_id, folio, nombre_archivo, hoja, n_fila))
            else:
                vistos[clave] = n_fila

            if muni_id is None:
                pendientes.append((
                    "MUNICIPIO_NO_RESUELTO", "ADVERTENCIA", "beneficiario_curp",
                    f"Fila {n_fila} de {nombre_archivo}: el municipio «{literal_muni}» no está "
                    "en el catálogo ni en los alias.",
                    "Agregar el alias en analitica.municipio_alias y recargar.",
                    anio, None, prog_id, folio, nombre_archivo, hoja, n_fila))
            if not an["curp_valida"]:
                pendientes.append((
                    "CURP_INVALIDA", "ADVERTENCIA", "beneficiario_curp",
                    f"Fila {n_fila} de {nombre_archivo} (folio {folio}): la CURP no pasó la "
                    f"validación ({an['motivo_invalidez']}); género, fecha de nacimiento y "
                    "rango de edad quedan vacíos.",
                    "Corregir la CURP en el padrón de origen y recargar. Prohibido inferir "
                    "sexo o edad por otros medios (R6).",
                    anio, muni_id, prog_id, folio, nombre_archivo, hoja, n_fila))

            if muni_id is None or prog_id is None:
                continue
            a = agregado[(anio, prog_id, muni_id)]
            a["apoyos"] += 1
            a["federal"] += federal
            a["estatal"] += estatal or 0.0
            a["municipal"] += municipal
            a["beneficiario"] += productor
            a["total"] += total or 0.0
            a["municipio_usado"] = muni_usado
            a["fuente_municipio"] = fuente_muni
            a["confianza"] = conf
            a["fuente_archivo"] = nombre_archivo
            a["fuente_hoja"] = hoja
        wb.close()
    return registros, agregado, c, pendientes


def cargar(archivo=None, dry_run=False):
    mapeo = comun.cargar_mapeo(MAPEO)
    rutas = _archivos(mapeo, archivo)
    registros, agregado, c, pendientes = leer(rutas, mapeo)

    observ = ("Agregado derivado del padrón folio por folio de Sequía 2023 "
              "(hojas consolidadas por región 23_Emer_Sequía_<REG>). Federal, municipal y "
              "aportación del productor en $0.00 por autorización expresa del " + FUENTE_DICTAMEN
              + ", no por imputación.")
    filas_apoyo = [
        (anio, prog_id, muni_id, a["apoyos"], a["federal"], a["estatal"], a["municipal"],
         a["beneficiario"], a["total"], a["fuente_archivo"], a["fuente_hoja"],
         a["municipio_usado"], a["fuente_municipio"], a["confianza"], "FOLIO", observ)
        for (anio, prog_id, muni_id), a in sorted(agregado.items())
    ]

    total_estatal = sum(a["estatal"] for a in agregado.values())
    total_apoyos = sum(a["apoyos"] for a in agregado.values())
    print(f"\nFolios leídos            : {c.leidas}")
    print(f"Municipios con agregado  : {len(agregado)}")
    print(f"Apoyos agregados         : {total_apoyos}")
    print(f"Apoyo estatal del padrón : ${total_estatal:,.2f}")
    print(f"Monto autorizado dictamen: ${MONTO_AUTORIZADO:,.2f}  ({FUENTE_DICTAMEN})")
    dif = MONTO_AUTORIZADO - total_estatal
    print(f"Diferencia (autorizado - ejercido): ${dif:,.2f}")

    if not dry_run:
        antes = db.contar("SELECT count(*)::int FROM analitica.beneficiario_curp")
        db.ejecutar_muchos(_SQL_CURP, registros)
        despues = db.contar("SELECT count(*)::int FROM analitica.beneficiario_curp")
        c.insertadas = despues - antes
        c.actualizadas = len(registros) - c.insertadas
        db.ejecutar_muchos(_SQL_APOYO, filas_apoyo)

        for (tipo, sev, entidad, desc, accion, anio, muni_id, prog_id,
             _folio, farch, fhoja, nfila) in pendientes:
            incidencias.registrar(tipo, sev, entidad, descripcion=desc, accion_sugerida=accion,
                                  anio=anio, municipio_id=muni_id, programa_id=prog_id,
                                  fuente_archivo=farch, fuente_hoja=fhoja, fila_origen=nfila)
            c.incidencias += 1

        if abs(dif) > 1.00:
            incidencias.registrar(
                "SUMA_APORTACIONES_NO_CUADRA", "ADVERTENCIA", "sequia_2023_dictamen",
                descripcion=(
                    f"El padrón folio por folio de Sequía 2023 suma ${total_estatal:,.2f} de "
                    f"apoyo estatal dictaminado sobre ${MONTO_AUTORIZADO:,.2f} autorizados; "
                    f"quedan ${dif:,.2f} sin ejercer respecto del oficio de autorización."),
                accion_sugerida=(
                    "No se fuerza el cuadre. Confirmar con la Dirección de Planeación si la "
                    "diferencia corresponde a subejercicio, a folios cancelados o a una ampliación "
                    "posterior, y documentarlo antes de usar la cifra en Glosa."),
                anio=2023, valor_origen=f"{total_estatal:.2f}",
                fuente_archivo="; ".join(os.path.basename(r) for r in rutas),
                fuente_hoja="23_Emer_Sequía_<REG>")
            c.incidencias += 1
    else:
        c.omitidas += len(registros)

    c.imprimir(dry_run)
    return 0


def main(argv=None):
    ap = argparse.ArgumentParser(
        description="Carga el padrón por folio del Programa Institucional Emergente por Sequía 2023.")
    ap.add_argument("--archivo", help="un solo .xlsx (por defecto: las 4 hojas regionales)")
    ap.add_argument("--dry-run", action="store_true", help="no escribe nada en la base")
    args = ap.parse_args(argv)
    try:
        return cargar(archivo=args.archivo, dry_run=args.dry_run)
    except FileNotFoundError as e:
        print(f"ERROR: {e}", file=sys.stderr)
        return 4
    except db.SinDatos as e:
        print(f"ERROR: no hay conexión a la base ({e}).", file=sys.stderr)
        return 2


if __name__ == "__main__":
    sys.exit(main())
