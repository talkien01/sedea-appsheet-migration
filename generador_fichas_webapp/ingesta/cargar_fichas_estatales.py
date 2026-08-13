"""Lee las fichas estatales por cadena productiva (apicultores, azúcar).

    python -m ingesta.cargar_fichas_estatales --dry-run
    python -m ingesta.cargar_fichas_estatales --archivo Ficha_Estatal_Apicultores_2026_AJUSTADA.xlsx

Estos archivos traen **municipio, beneficiarios y monto**, sin desglose
federal/municipal y sin programa_id explícito. Se leen, se reportan y se
contrastan; las dos versiones de la ficha de azúcar no coinciden entre sí, y esa
diferencia se registra como incidencia en vez de elegir una en silencio (R8).
"""
import argparse
import sys

import openpyxl

import db
from ingesta import comun
from services import incidencias

MAPEO = "fichas_estatales.json"


def leer_ficha(archivo, cfg):
    ruta = comun.ruta_archivo(archivo)
    wb = openpyxl.load_workbook(ruta, data_only=True)
    hoja = cfg["hoja"] if cfg["hoja"] in wb.sheetnames else wb.sheetnames[0]
    ws = wb[hoja]
    filas = list(ws.iter_rows(values_only=True))
    f_enc = cfg["fila_encabezado"] - 1
    encabezados = list(filas[f_enc])
    if cfg.get("fila_subencabezado"):
        sub = filas[cfg["fila_subencabezado"] - 1]
        grupo = None
        combinados = []
        for i, h in enumerate(encabezados):
            if h not in (None, ""):
                grupo = str(h).strip()
            s = sub[i] if i < len(sub) else None
            combinados.append(f"{grupo} {s}".strip() if s else grupo)
        encabezados = combinados
    datos = []
    for n, fila in enumerate(filas[cfg["fila_datos_desde"] - 1:], start=cfg["fila_datos_desde"]):
        if not fila or not fila[0]:
            continue
        etiqueta = str(fila[0]).strip()
        if comun.normalizar(etiqueta) in ("TOTAL", "SUBTOTAL", "SUMA"):
            continue
        datos.append({"fila": n, "municipio": etiqueta,
                      "valores": {encabezados[i]: fila[i] for i in range(1, min(len(encabezados), len(fila)))}})
    return hoja, encabezados, datos


def main(argv=None):
    ap = argparse.ArgumentParser(description="Lectura de las fichas estatales por cadena productiva.")
    ap.add_argument("--archivo", help="solo este archivo; por defecto, todos los del mapeo")
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args(argv)

    mapeo = comun.cargar_mapeo(MAPEO)
    archivos = [args.archivo] if args.archivo else list(mapeo["archivos"])
    c = comun.Contadores()
    montos_por_archivo = {}

    for archivo in archivos:
        cfg = mapeo["archivos"].get(archivo)
        if not cfg:
            print(f"ERROR: {archivo} no está en {MAPEO}.", file=sys.stderr)
            return 3
        try:
            hoja, encabezados, datos = leer_ficha(archivo, cfg)
        except FileNotFoundError as e:
            print(f"ERROR: {e}", file=sys.stderr)
            return 4
        print(f"\nArchivo : {archivo}")
        print(f"Hoja    : {hoja}")
        print("Encabezados detectados:", [h for h in encabezados if h])
        print(f"Municipios con dato: {len(datos)}")
        c.leidas += len(datos)
        total = 0.0
        for d in datos:
            valores = {comun.normalizar(k): comun.a_numero(v) for k, v in d["valores"].items()}
            monto = (valores.get("TOTAL MONTO") or valores.get("MONTO"))
            if monto is not None:
                total += monto
        montos_por_archivo[archivo] = total
        print(f"Monto total del archivo: {total:,.2f}")
        c.omitidas += len(datos)  # lectura de referencia: no se escriben importes

    azucar = [a for a in montos_por_archivo if "Azucar" in a or "Azúcar" in a]
    if len(azucar) > 1:
        a1, a2 = azucar[0], azucar[1]
        if abs(montos_por_archivo[a1] - montos_por_archivo[a2]) > 1.00:
            print(f"\nDIFERENCIA entre versiones de la ficha de azúcar: "
                  f"{a1}={montos_por_archivo[a1]:,.2f} vs {a2}={montos_por_archivo[a2]:,.2f}")
            if not args.dry_run:
                try:
                    incidencias.registrar(
                        "FUENTE_FALTANTE", "ADVERTENCIA", "fichas_estatales",
                        descripcion=(f"Las dos versiones de la ficha estatal de azúcar no coinciden: "
                                     f"{a1}={montos_por_archivo[a1]:.2f} vs {a2}={montos_por_archivo[a2]:.2f}."),
                        accion_sugerida=("Definir con el área cuál versión es la oficial antes de usar "
                                         "estas cifras en la Glosa; no se promedian ni se elige en silencio."),
                        anio=2026, valor_origen=f"{a1}|{a2}", fuente_archivo=a2, fuente_hoja="ESTATAL")
                    c.incidencias += 1
                except db.SinDatos:
                    print("(sin base: la incidencia no se pudo registrar)")
    print()
    c.imprimir(args.dry_run)
    return 0


if __name__ == "__main__":
    sys.exit(main())
