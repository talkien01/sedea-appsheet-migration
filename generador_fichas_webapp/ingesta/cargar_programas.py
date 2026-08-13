"""Da de alta programas y alias faltantes a partir de los componentes reales.

    python -m ingesta.cargar_programas --dry-run
    python -m ingesta.cargar_programas

Fuentes: los COMPONENTE de `Base Cadereyta Programas sedea.xlsx` y los
componentes de `analitica.v_oficial_municipio` (2026). Solo se crea lo que
existe en el origen; no se inventan programas.
"""
import argparse
import sys

import openpyxl

import db
from ingesta import comun

ARCHIVO = "Base Cadereyta Programas sedea.xlsx"
HOJA = "Hoja1"

# Equivalencias verificadas contra el catálogo real de analitica.programa.
ALIAS_CONOCIDOS = {
    "TECNIFICACIÓN DEL RIEGO": "TECNIFICACIÓN DE RIEGO",
}


def componentes_del_xlsx():
    ruta = comun.ruta_archivo(ARCHIVO)
    wb = openpyxl.load_workbook(ruta, data_only=True)
    ws = wb[HOJA] if HOJA in wb.sheetnames else wb.worksheets[0]
    filas = list(ws.iter_rows(values_only=True))
    indice = comun.indexar_encabezados(filas[0])
    col = comun.resolver_columna(indice, ["COMPONENTE", "PROGRAMA"])
    if col is None:
        return []
    return sorted({str(f[col]).strip() for f in filas[1:] if f[col]})


def componentes_oficiales():
    return [r["componente"] for r in db.consultar(
        "SELECT DISTINCT componente FROM analitica.v_oficial_municipio "
        "WHERE componente IS NOT NULL ORDER BY 1")]


def main(argv=None):
    ap = argparse.ArgumentParser(description="Alta de programas y alias desde los componentes reales.")
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args(argv)

    try:
        por_nombre, alias = comun.catalogo_programas(db)
        c = comun.Contadores()
        origenes = [(ARCHIVO, HOJA, componentes_del_xlsx()),
                    ("analitica.v_oficial_municipio", "vista", componentes_oficiales())]
        for archivo, hoja, componentes in origenes:
            print(f"{archivo} / {hoja}: {len(componentes)} componente(s) distinto(s)")
            for nombre in componentes:
                c.leidas += 1
                n = comun.normalizar(nombre)
                if n in por_nombre or n in alias:
                    c.omitidas += 1
                    continue
                destino = ALIAS_CONOCIDOS.get(nombre)
                if destino and comun.normalizar(destino) in por_nombre:
                    pid = por_nombre[comun.normalizar(destino)]
                    print(f"  alias  «{nombre}» -> {destino} (programa_id={pid})")
                    if not args.dry_run:
                        db.ejecutar("INSERT INTO analitica.programa_alias (alias, programa_id) "
                                    "VALUES (%s,%s) ON CONFLICT DO NOTHING", (nombre, pid))
                    alias[n] = pid
                    c.actualizadas += 1
                    continue
                print(f"  alta   «{nombre}» (no existía en analitica.programa)")
                if not args.dry_run:
                    db.ejecutar("INSERT INTO analitica.programa (nombre) VALUES (%s) "
                                "ON CONFLICT DO NOTHING", (nombre,))
                    pid = db.escalar("SELECT programa_id FROM analitica.programa WHERE nombre=%s",
                                     (nombre,))
                    if pid:
                        por_nombre[n] = int(pid)
                c.insertadas += 1
        c.imprimir(args.dry_run)
        if c.insertadas and not args.dry_run:
            print("Recuerda reclasificar: python -m services.clasificacion --aplicar")
    except db.SinDatos as e:
        print(f"ERROR: sin base de datos ({e}).", file=sys.stderr)
        return 2
    except FileNotFoundError as e:
        print(f"ERROR: {e}", file=sys.stderr)
        return 4
    return 0


if __name__ == "__main__":
    sys.exit(main())
