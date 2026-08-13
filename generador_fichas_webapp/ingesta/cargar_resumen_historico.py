"""Lee el machote 'Resumen Histórico por Municipio' y reporta qué trae de verdad.

    python -m ingesta.cargar_resumen_historico --archivo "Resumen Histórico por Municipio.xlsx" --dry-run
    python -m ingesta.cargar_resumen_historico --archivo "Resumen_Historico_por_Municipio_llenado_2026_Cadereyta.xlsx" --anio 2026

Hallazgo verificado al abrir los archivos (2026-08-13):

* `Resumen Histórico por Municipio.xlsx` es la **plantilla en blanco**: todas las
  celdas de datos están vacías. No hay 2021 ni 2022 que cargar. En vez de
  inventarlos se registra la incidencia `ANIO_SIN_DATOS` (R1/R8).
* La versión llenada solo tiene 2026 de la región Cadereyta, con grano
  municipio × clasificación (Emergentes/Productividad) — **sin programa**. Como
  2026 ya está cubierto por la fuente autorizada `v_oficial_municipio` (A5),
  cargarlo otra vez en `apoyo_municipio` duplicaría la matriz. Por eso este
  script lo usa como **contraste**: compara sus totales contra la base y levanta
  incidencia si difieren, sin escribir importes.
"""
import argparse
import sys

import openpyxl

import db
from ingesta import comun
from services import incidencias

MAPEO = "resumen_historico_municipio.json"


def _leer_emer_prod(ws, mapeo):
    cfg = mapeo["hojas"]["Por Emer-Prod"]
    filas = list(ws.iter_rows(values_only=True))
    f_anios = filas[cfg["fila_anios"] - 1]
    f_sub = filas[cfg["fila_subcolumna"] - 1]
    f_metrica = filas[cfg["fila_metrica"] - 1]

    # El encabezado de año viene combinado: se arrastra hacia la derecha.
    anios, actual = [], None
    for v in f_anios:
        if v not in (None, ""):
            actual = comun.a_entero(v)
        anios.append(actual)
    metricas, actual = [], None
    for v in f_metrica:
        if v not in (None, ""):
            actual = str(v).strip().upper()
        metricas.append(actual)

    omitir = {comun.normalizar(x) for x in mapeo["filas_agregadas_a_omitir"]}
    datos = []
    for n_fila, fila in enumerate(filas[cfg["fila_datos_desde"] - 1:], start=cfg["fila_datos_desde"]):
        etiqueta = fila[cfg["columna_municipio"] - 1]
        if not etiqueta or comun.normalizar(etiqueta) in omitir:
            continue
        for i, celda in enumerate(fila):
            if i == cfg["columna_municipio"] - 1 or celda in (None, ""):
                continue
            valor = comun.a_numero(celda)
            if valor is None:
                continue
            datos.append({
                "fila": n_fila,
                "municipio": str(etiqueta).strip(),
                "anio": anios[i] if i < len(anios) else None,
                "metrica": metricas[i] if i < len(metricas) else None,
                "subcolumna": (str(f_sub[i]).strip().upper() if i < len(f_sub) and f_sub[i] else None),
                "valor": valor,
            })
    return datos


def main(argv=None):
    ap = argparse.ArgumentParser(description="Lectura y contraste del Resumen Histórico por Municipio.")
    ap.add_argument("--archivo", required=True)
    ap.add_argument("--anio", type=int, help="ejercicio a contrastar")
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args(argv)

    try:
        mapeo = comun.cargar_mapeo(MAPEO)
        ruta = comun.ruta_archivo(args.archivo)
        wb = openpyxl.load_workbook(ruta, data_only=True)
        hoja = "Por Emer-Prod"
        if hoja not in wb.sheetnames:
            print(f"ERROR: el archivo no tiene la hoja «{hoja}».", file=sys.stderr)
            return 3
        print(f"Archivo : {args.archivo}")
        print(f"Hoja    : {hoja}")
        datos = _leer_emer_prod(wb[hoja], mapeo)
        c = comun.Contadores()
        c.leidas = len(datos)
        if not datos:
            print("La hoja está VACÍA: es la plantilla en blanco, no hay nada que cargar.")
        anios = sorted({d["anio"] for d in datos if d["anio"]})
        print("Años con algún dato en el archivo:", anios or "ninguno")

        por_nombre, alias, pseudo = comun.catalogo_municipios(db)
        objetivo = args.anio
        contrastadas = 0
        for d in datos:
            if objetivo and d["anio"] != objetivo:
                c.omitidas += 1
                continue
            if d["subcolumna"] != "SUBTOTAL" or d["metrica"] != "MONTO":
                c.omitidas += 1
                continue
            muni_id, _f, _c = comun.resolver_municipio(d["municipio"], por_nombre, alias, pseudo)
            if muni_id is None:
                c.omitidas += 1
                continue
            en_base = db.escalar(
                "SELECT sum(total) FROM analitica.vw_matriz_historica WHERE anio=%s AND municipio_id=%s",
                (d["anio"], muni_id))
            en_base = float(en_base) if en_base is not None else None
            contrastadas += 1
            if en_base is None or abs(en_base - d["valor"]) > 1.00:
                print(f"  DIFERENCIA {d['municipio']} {d['anio']}: archivo={d['valor']} base={en_base}")
                if not args.dry_run:
                    incidencias.registrar(
                        "SUMA_APORTACIONES_NO_CUADRA", "ADVERTENCIA", "resumen_historico_xlsx",
                        descripcion=(f"{d['municipio']} {d['anio']}: el machote reporta "
                                     f"{d['valor']} y la base {en_base}."),
                        accion_sugerida=("Conciliar el machote con analitica antes de publicar; "
                                         "no se sobreescribe ninguno de los dos automáticamente."),
                        anio=d["anio"], municipio_id=muni_id, valor_origen=str(d["valor"]),
                        fuente_archivo=args.archivo, fuente_hoja=hoja, fila_origen=d["fila"])
                    c.incidencias += 1
            else:
                c.actualizadas += 1
        print(f"Filas contrastadas contra la base: {contrastadas}")

        faltantes = [a for a in (2021, 2022) if a not in anios]
        for a in faltantes:
            existe = db.escalar(
                "SELECT count(*)::int FROM analitica.apoyo_municipio WHERE anio=%s", (a,), default=0)
            if int(existe or 0) == 0 and not args.dry_run:
                incidencias.registrar(
                    "ANIO_SIN_DATOS", "ADVERTENCIA", "apoyo_municipio",
                    descripcion=(f"El ejercicio {a} no está en apoyo_municipio y el machote "
                                 f"«{args.archivo}» tampoco lo trae lleno."),
                    accion_sugerida=(f"Solicitar al área de datos el desglose municipal {a}. "
                                     f"Mientras tanto ese año se rinde vacío, nunca como cero."),
                    anio=a, fuente_archivo=args.archivo, fuente_hoja=hoja)
                c.incidencias += 1
        c.imprimir(args.dry_run)
    except FileNotFoundError as e:
        print(f"ERROR: {e}", file=sys.stderr)
        return 4
    except db.SinDatos as e:
        print(f"ERROR: sin base de datos ({e}).", file=sys.stderr)
        return 2
    return 0


if __name__ == "__main__":
    sys.exit(main())
