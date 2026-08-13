import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Mapa de disponibilidad"

FONT = "Arial"
header_fill = PatternFill("solid", fgColor="366092")
header_font = Font(name=FONT, bold=True, color="FFFFFF", size=11)
group_fill = PatternFill("solid", fgColor="D9E1F2")
group_font = Font(name=FONT, bold=True, size=10)
normal_font = Font(name=FONT, size=10)
ok_fill = PatternFill("solid", fgColor="C6EFCE")
partial_fill = PatternFill("solid", fgColor="FFEB9C")
missing_fill = PatternFill("solid", fgColor="FFC7CE")
thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

cols = ["Rubro de la ficha", "Nivel", "Disponible en Drive", "Archivo / hoja fuente", "Estado", "Notas"]
ws.append(cols)
for c in range(1, len(cols)+1):
    cell = ws.cell(row=1, column=c)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(wrap_text=True, vertical="center")
    cell.border = border
ws.freeze_panes = "A2"

rows = [
 # (rubro, nivel, disponible?, fuente, estado, notas)
 ("Resumen de apoyos por tipo (Productivo/Emergente) y aportación F/E/M/Productores","Estatal","Sí","Carpetas Chucho 2025 e Historicos > Copia de ESTATAL (hoja Avance Estatal 2025) + GLOSA 2025>Fichas>Tablas origen>2025 0900 INVERSIONES GLOBAL actualizado.xlsx","Disponible","Cifras a 2025; falta confirmar corte 2026 (03-ago) usado en el machote."),
 ("Resumen de apoyos por tipo (Productivo/Emergente) y aportación F/E/M/Productores","Regional SJR","Sí","Carpetas Chucho 2025 e Historicos > Eventos _REGIÓN SJR_ > 2025 0904 REG SJR.xlsx (hojas Dinamismo2025Estatal, Municipalizado2025Estatal, Riesgos 2024, etc.)","Disponible","Requiere consolidar varias hojas por concepto."),
 ("Resumen de apoyos por tipo (Productivo/Emergente) y aportación F/E/M/Productores","Amealco","Sí","Eventos AMEALC_ > 2026 0309 AMEALC_.xlsx + 2026 0309 AMEALC.docx (ficha ya llenada)","Disponible","Es la ficha de referencia que compartió el usuario."),
 ("Resumen de apoyos por tipo (Productivo/Emergente) y aportación F/E/M/Productores","Pedro Escobedo","Parcial","2025 0904 REG SJR.xlsx: filas 'PEDRO ESCOBEDO' dentro de Municipalizado23/24/25, Dinamismo23/24/25, MaízBlanco2025, Bordería2025, Tecnificación2025","Parcial","Datos dispersos en varias hojas de la región; no existe una hoja única 'Pedro Escobedo' como sí existe 'Amealco'. Hay que consolidarla (mismo patrón que la hoja Amealco)."),
 ("Resumen de apoyos por tipo (Productivo/Emergente) y aportación F/E/M/Productores","Otros municipios / regiones (16 municipios, 6 regionales restantes)","No / desconocido","El usuario está subiendo carpetas nuevas a 'Carpetas Chucho 2025 e Historicos'","Pendiente","Región QRO tiene un archivo (20241003 REGIÓN QRO.xlsx) pero desactualizado (oct-2024)."),

 ("Inversión total 2021-2026 por año (apoyos, beneficiarios, aportaciones, total)","Estatal","Sí","Copia de ESTATAL (hojas TOTAL 2021...TOTAL 2025) + GLOSA 2025>Tablas origen","Disponible","Serie histórica completa 2021-2025; 2026 parcial (en curso)."),
 ("Inversión total 2021-2026 por año (apoyos, beneficiarios, aportaciones, total)","Regional SJR","Parcial","2025 0904 REG SJR.xlsx (Inversiones Admón, GestiónDeRiesgos, Hoja1)","Parcial","Serie por año existe pero repartida en varias hojas."),
 ("Inversión total 2021-2026 por año (apoyos, beneficiarios, aportaciones, total)","Amealco","Sí","2026 0309 AMEALC_.xlsx (hoja 'Amealco', serie 2022-2025 por concepto y año)","Disponible","Coincide con Tabla 1 de la ficha de referencia."),
 ("Inversión total 2021-2026 por año (apoyos, beneficiarios, aportaciones, total)","Pedro Escobedo","No","No se localizó una hoja equivalente a 'Amealco' para Pedro Escobedo","No localizado","Se puede construir agregando las filas de Pedro Escobedo de cada hoja anual de la región, pero falta el desglose fino (H/M, productivo vs emergente) que sí tiene Amealco."),

 ("Extensión territorial y superficie agrícola (riego/temporal)","Estatal","Sí","Copia de ESTATAL (encabezado 'CIFRAS ESTADISTICAS ESTADO DE QUERÉTARO')","Disponible",""),
 ("Extensión territorial y superficie agrícola (riego/temporal)","Amealco","Sí","Ficha de referencia (F0159-F0163) — fuente original probablemente INEGI / SIAP, no localizada aún en Drive como tabla editable","Disponible en la ficha","Vale la pena pedir la fuente INEGI/SIAP para poder replicar en los demás municipios."),
 ("Extensión territorial y superficie agrícola (riego/temporal)","Pedro Escobedo y resto de municipios","No","No localizado","No localizado","Requiere fuente municipal tipo SIAP/INEGI por municipio; no vista aún en el Drive."),

 ("Top 5-6 productos (superficie, volumen, valor) por nivel","Estatal","Sí","Copia de ESTATAL (RANKING NACIONAL, Productos destacados)","Disponible",""),
 ("Top 5-6 productos (superficie, volumen, valor) por nivel","Amealco","Sí","Ficha de referencia (F0164-F0181)","Disponible en la ficha","Fuente base (SIAP) no confirmada en Drive."),
 ("Top 5-6 productos (superficie, volumen, valor) por nivel","Pedro Escobedo y resto","No","No localizado","No localizado","Mismo comentario que extensión territorial."),

 ("Precipitación mensual por año (histórico ~2012-2026)","Estatal","Sí","GLOSA 2025 > Tablas origen > Precipitación_alm_sequia.xlsx; PDF 'Precipitación Mensual CONAGUA 2012-2025'; y archivos diarios 'PRESAS/PRECIPITACION [fecha] 2026.xlsx' (carpeta compartida, actualización por día)","Disponible","Esta serie es estatal; no está desagregada por municipio en lo que se revisó."),
 ("Precipitación mensual por año (histórico ~2012-2026)","Amealco / Pedro Escobedo / municipios","Parcial","Mismos archivos CONAGUA, si tienen estación por municipio","Por confirmar","Falta revisar si el archivo CONAGUA trae desagregación por estación/municipio o solo estatal."),

 ("Distribución de beneficiarios por sexo y grupo de edad","Estatal","Sí","Copia de ESTATAL / Hoja 2 (columnas H, M por concepto)","Disponible",""),
 ("Distribución de beneficiarios por sexo y grupo de edad","Regional SJR","Sí","Ficha de referencia (F0488-F0511) + 2025 0904 REG SJR.xlsx Hoja 2","Disponible",""),
 ("Distribución de beneficiarios por sexo y grupo de edad","Amealco","Sí","Ficha de referencia (F0512-F0536) + hoja 'Amealco' (columnas H, M)","Disponible",""),
 ("Distribución de beneficiarios por sexo y grupo de edad","Pedro Escobedo","Parcial","Columnas H/M existen en algunas filas de 2025 0904 REG SJR.xlsx pero incompletas para Pedro Escobedo","Parcial","Muchas celdas H/M vacías o en blanco para Pedro Escobedo en el archivo regional."),

 ("Situación de almacenamientos / presas","Estatal","Sí","GLOSA 2025 > Tablas origen > PRESAS ALMACENAMIENTOS HOY ACTUALIZADO.pdf + archivos diarios 'PRESAS [fecha] 2026.xlsx'","Disponible","Es información estatal, no por municipio."),

 ("Programa institucional 'Apoyo al Campo Queretano' 2026 — entrega parcial por concepto","Estatal","Sí","Ficha de referencia (F0402-F0461) — fuente probablemente el mismo padrón/base que alimenta 'Copia de ESTATAL'","Por confirmar origen exacto","No se localizó una hoja 2026 equivalente a 'Copia de ESTATAL' pero con corte 03-ago; puede ser la misma base con filtro de fecha."),
 ("Programa institucional 'Apoyo al Campo Queretano' 2026 — entrega parcial por concepto","Regional SJR / Amealco","Sí","Ficha de referencia (F0412-F0461, F0462-F0536)","Disponible en la ficha",""),
 ("Programa institucional 'Apoyo al Campo Queretano' 2026 — entrega parcial por concepto","Pedro Escobedo y resto","No","No localizado","No localizado","Mismo dato pero desagregado por municipio; no se ha ubicado la base 2026 en Drive todavía."),

 ("Metodología / definición de rubros","Todos","Sí","Doc 'Fichas SEDEA — Rubros y Consultas SQL (Región SJR)' — describe una base Postgres local (esquema 'analitica') con vistas por programa/año/municipio","Fuente primaria = base de datos, no Drive","La fuente 'oficial' de estos datos es una base de datos Postgres local, no el Drive. Los xlsx del Drive son extracciones/curaciones manuales de esa base. Si el equipo tiene acceso a esa base, sería la fuente más confiable y automatizable."),
]

for r in rows:
    ws.append(list(r))

# style data rows
last_row = ws.max_row
status_col = 5
for row in range(2, last_row+1):
    for col in range(1, len(cols)+1):
        cell = ws.cell(row=row, column=col)
        cell.font = normal_font
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.border = border
    status_val = str(ws.cell(row=row, column=status_col).value or "")
    fill = None
    if "Disponible" in status_val and "No" not in status_val:
        fill = ok_fill
    elif "Parcial" in status_val or "confirmar" in status_val or "Por confirmar" in status_val:
        fill = partial_fill
    elif "No localizado" in status_val or "Pendiente" in status_val:
        fill = missing_fill
    if fill:
        ws.cell(row=row, column=status_col).fill = fill

widths = [42, 20, 14, 55, 20, 55]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.row_dimensions[1].height = 34

# Second sheet: raw Pedro Escobedo figures extracted, for traceability
ws2 = wb.create_sheet("Pedro Escobedo - datos crudos")
ws2.append(["Hoja fuente (2025 0904 REG SJR.xlsx)", "Concepto", "Apoyos/Beneficiarios", "Apoyo Federal", "Apoyo Estatal", "Apoyo Municipal", "Aportación Productores", "Total", "% part. estatal"])
for c in range(1,10):
    cell = ws2.cell(row=1, column=c)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(wrap_text=True, vertical="center")

pe_rows = [
 ("Hoja 1 (acumulado región, corte sept-2025)","Total apoyos 2022-2026 (a la fecha del archivo)",4769,9240144,68931547,8975000,63772100,150918791,""),
 ("RESUMEN DE INVERSIONES REGIÓN SJR 2022-2024 ACUMULADO","Acumulado 2022-2024",3703,9240144,55781255,6000000,46893430,117914830,""),
 ("Municipalizado23Estatal","Municipalizado 2023",117,"",3000000,3000000,5072759,11072759,"7.7%"),
 ("Municipalizado2025Estatal (región SJR)","Municipalizado 2025",245,"",2765993,2765993,2765993,8297979,"28.4%"),
 ("Dinamismo2023Estatal","Dinamismo 2023",36,"",2158569,"",3035276,5193845,"9%"),
 ("Dinamismo2024Estatal","Dinamismo 2024",159,"",5635835,"",11947960,17583795,"10%"),
 ("Dinamismo2025Estatal (región SJR)","Dinamismo 2025",185,"",7103737,"",17017360,24121097,"21.7%"),
 ("Maíz Blanco2025Estatal","Maíz Blanco 2025 (636 apoyos, 199.7 ton, 1,391 ha)",636,"",1705310,"","","","9.8%"),
 ("Pacas&Suplementos2025","Pacas y suplementos 2025 (62.7 ton pacas)",72,"",300888,"","","","20.1%"),
 ("Bordería2025Estatal","Bordería 2025",2,"",160685,"",93790,254475,"1.4%"),
 ("Tecnificación2025Estatal","Tecnificación de riego 2025",33,"",2800637,"",1716368,4517005,"24.4%"),
]
for r in pe_rows:
    ws2.append(list(r))
for row in range(2, ws2.max_row+1):
    for col in range(1,10):
        cell = ws2.cell(row=row, column=col)
        cell.font = normal_font
        cell.border = border
        cell.alignment = Alignment(wrap_text=True, vertical="top")
w2 = [30,34,16,14,14,14,16,14,12]
for i,w in enumerate(w2, start=1):
    ws2.column_dimensions[get_column_letter(i)].width = w
ws2.freeze_panes = "A2"

note_row = ws2.max_row + 2
ws2.cell(row=note_row, column=1, value="Nota: cifras extraídas manualmente de 2025 0904 REG SJR.xlsx (Drive, programas.sedea@queretaro.gob.mx). No incluye desglose por sexo/edad, extensión territorial, top de productos ni precipitación específicos de Pedro Escobedo, que no se localizaron en el Drive para este municipio.").font = Font(name=FONT, italic=True, size=9)

wb.save("/sessions/vibrant-intelligent-goldberg/mnt/outputs/work/Mapa_disponibilidad_datos_fichas.xlsx")
print("saved")
