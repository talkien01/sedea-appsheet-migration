import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FONT="Arial"
header_fill = PatternFill("solid", fgColor="366092")
header_font = Font(name=FONT, bold=True, color="FFFFFF", size=11)
input_font = Font(name=FONT, color="0000FF", size=10)   # blue = celda a llenar
example_fill = PatternFill("solid", fgColor="FFF2CC")
normal_font = Font(name=FONT, size=10)
thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

municipios = ["AMEALCO DE BONFIL","ARROYO SECO","CADEREYTA DE MONTES","COLÓN","CORREGIDORA",
 "EL MARQUÉS","EZEQUIEL MONTES","HUIMILPAN","JALPAN DE SERRA","LANDA DE MATAMOROS",
 "PEDRO ESCOBEDO","PEÑAMILLER","PINAL DE AMOLES","QUERÉTARO","SAN JOAQUÍN",
 "SAN JUAN DEL RÍO","TEQUISQUIAPAN","TOLIMÁN"]

wb = openpyxl.Workbook()
wb.remove(wb.active)

def header(ws, cols, height=32):
    ws.append(cols)
    for c in range(1, len(cols)+1):
        cell = ws.cell(row=1, column=c)
        cell.font = header_font; cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.border = border
    ws.row_dimensions[1].height = height
    ws.freeze_panes = "A2"

def style_body(ws, ncols, example_rows=1):
    for row in range(2, ws.max_row+1):
        is_example = (row <= 1+example_rows)
        for col in range(1, ncols+1):
            cell = ws.cell(row=row, column=col)
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if is_example:
                cell.font = normal_font
                cell.fill = example_fill
            else:
                cell.font = input_font

# ---- Hoja 0: LEEME ----
ws0 = wb.create_sheet("LEEME")
ws0.append(["Plantilla de datos de referencia manual — 4 bloques que no viven en Postgres"])
ws0["A1"].font = Font(name=FONT, bold=True, size=14)
lines = [
 "",
 "Uso: llenar una fila por municipio en cada hoja (Territorio, Productos_top, Precipitacion, Demografia_municipal).",
 "La primera fila de datos (fondo amarillo) es un EJEMPLO con los valores reales de Amealco de Bonfil, tomados de la ficha que compartió el equipo — sirve de referencia de formato, no hay que borrarla.",
 "Las celdas en AZUL son las que hay que llenar; el resto (encabezados) no se toca.",
 "Fuente sugerida por bloque:",
 "  - Territorio y superficie agrícola: INEGI / SIAP (marco geoestadístico + cierre agrícola).",
 "  - Top de productos: SIAP, Cierre de la Producción Agrícola y Pecuaria por municipio.",
 "  - Precipitación: CONAGUA (estación meteorológica más cercana al municipio) o el PDF 'Precipitación Mensual CONAGUA 2012-2025' ya en el Drive, si trae desagregación.",
 "  - Demografía municipal (sexo/edad de beneficiarios): no hay fuente confirmada todavía — pendiente de que el equipo decida de dónde sacarlo (padrón de beneficiarios por municipio, quizás).",
 "",
 "Estos 4 bloques NO se generan automáticamente. Cuando se arme una ficha, el resto de la información (apoyos, inversión, avance) sale de Postgres; estos 4 bloques se copian de aquí.",
]
for l in lines:
    ws0.append([l])
for row in range(2, ws0.max_row+1):
    ws0.cell(row=row, column=1).font = Font(name=FONT, size=10)
    ws0.cell(row=row, column=1).alignment = Alignment(wrap_text=True)
ws0.column_dimensions["A"].width = 110

# ---- Hoja 1: Territorio ----
ws1 = wb.create_sheet("Territorio")
header(ws1, ["Municipio","Extensión territorial (Ha)","% del territorio estatal","Superficie agrícola total (Ha)","Riego (Ha)","Riego - Unidades de Producción","Temporal (Ha)","Temporal - Unidades de Producción","Fuente / fecha de corte"])
ws1.append(["AMEALCO DE BONFIL (ejemplo)",71333.3,"6.1%","(sumar riego+temporal)",4844,"(no en ficha)",13633,"(no en ficha)","Ficha 03-ago-2026 compartida por el equipo"])
for m in municipios:
    if m != "AMEALCO DE BONFIL":
        ws1.append([m,"","","","","","","",""])
style_body(ws1, 9)
widths1 = [24,16,14,16,10,14,10,14,30]
for i,w in enumerate(widths1, start=1): ws1.column_dimensions[get_column_letter(i)].width = w

# ---- Hoja 2: Productos_top ----
ws2 = wb.create_sheet("Productos_top")
header(ws2, ["Municipio","Rank","Producto","Superficie (Ha)","Volumen (Ton)","Valor (MDP)","Fuente / fecha de corte"])
ejemplo2 = [
 ["AMEALCO DE BONFIL (ejemplo)",1,"Maíz grano",14680,63718,316,"Ficha 03-ago-2026"],
 ["AMEALCO DE BONFIL (ejemplo)",2,"Hongos, setas y champiñones",8,3450,95,"Ficha 03-ago-2026"],
]
for r in ejemplo2: ws2.append(r)
for m in municipios:
    if m != "AMEALCO DE BONFIL":
        ws2.append([m,1,"","","",""," "])
        ws2.append([m,2,"","","",""," "])
style_body(ws2, 7, example_rows=2)
widths2 = [24,8,28,16,16,14,26]
for i,w in enumerate(widths2, start=1): ws2.column_dimensions[get_column_letter(i)].width = w

# ---- Hoja 3: Precipitacion ----
ws3 = wb.create_sheet("Precipitacion")
meses = ["ENE","FEB","MAR","ABR","MAY","JUN","JUL","AGO","SEP","OCT","NOV","DIC"]
header(ws3, ["Municipio","Año"]+meses+["Anual","Fuente / estación CONAGUA"])
ws3.append(["AMEALCO DE BONFIL (ejemplo)",2025,3.4,8.2,0.4,0.4,73.9,163.2,104.6,91.8,127.8,80.0,4.6,14.7,"CONAGUA — estación pendiente de confirmar"])
for m in municipios:
    if m != "AMEALCO DE BONFIL":
        ws3.append([m,2025]+[""]*12+[""])
style_body(ws3, 15)
w3 = [24,8]+[8]*12+[10,30]
for i,w in enumerate(w3, start=1): ws3.column_dimensions[get_column_letter(i)].width = w

# ---- Hoja 4: Demografia_municipal ----
ws4 = wb.create_sheet("Demografia_municipal")
header(ws4, ["Municipio","Grupo de edad","Hombres","Mujeres","Total","% del total municipal","Fuente / fecha de corte"])
ejemplo4 = [
 ["AMEALCO DE BONFIL (ejemplo)","Jóvenes (18-29)",12,7,19,"4.2%","Ficha 03-ago-2026"],
 ["AMEALCO DE BONFIL (ejemplo)","Adultos (30-59)",154,85,239,"52.8%","Ficha 03-ago-2026"],
 ["AMEALCO DE BONFIL (ejemplo)","Adultos mayores (60+)",152,43,195,"43.0%","Ficha 03-ago-2026"],
]
for r in ejemplo4: ws4.append(r)
for m in municipios:
    if m != "AMEALCO DE BONFIL":
        for grupo in ["Jóvenes (18-29)","Adultos (30-59)","Adultos mayores (60+)"]:
            ws4.append([m,grupo,"","","","",""])
style_body(ws4, 7, example_rows=3)
w4 = [24,20,10,10,10,14,26]
for i,w in enumerate(w4, start=1): ws4.column_dimensions[get_column_letter(i)].width = w

wb.save("/sessions/vibrant-intelligent-goldberg/mnt/outputs/work/Datos_referencia_manual_municipios.xlsx")
print("saved")
