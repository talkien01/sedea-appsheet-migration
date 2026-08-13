import csv, openpyxl
from collections import defaultdict
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE = "/sessions/vibrant-intelligent-goldberg/mnt/analitica_export/"
FONT = "Arial"
header_fill = PatternFill("solid", fgColor="366092")
header_font = Font(name=FONT, bold=True, color="FFFFFF", size=11)
normal_font = Font(name=FONT, size=10)
ok_fill = PatternFill("solid", fgColor="C6EFCE")
partial_fill = PatternFill("solid", fgColor="FFEB9C")
missing_fill = PatternFill("solid", fgColor="FFC7CE")
thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def style_header(ws, ncols, height=30):
    for c in range(1, ncols+1):
        cell = ws.cell(row=1, column=c)
        cell.font = header_font; cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.border = border
    ws.row_dimensions[1].height = height
    ws.freeze_panes = "A2"

def style_rows(ws, ncols, start=2):
    for row in range(start, ws.max_row+1):
        for col in range(1, ncols+1):
            cell = ws.cell(row=row, column=col)
            cell.font = normal_font
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = border

wb = openpyxl.load_workbook("/sessions/vibrant-intelligent-goldberg/mnt/outputs/work/Mapa_disponibilidad_datos_fichas.xlsx")

# ---- Update key rows on "Mapa de disponibilidad" re: demografia (now confirmed EMPTY) ----
ws = wb["Mapa de disponibilidad"]
for row in range(2, ws.max_row+1):
    rubro = ws.cell(row=row, column=1).value or ""
    nivel = ws.cell(row=row, column=2).value or ""
    if "Distribución de beneficiarios por sexo" in rubro:
        ws.cell(row=row, column=4, value="analitica.beneficiarios_demografia CONFIRMADA VACÍA (0 filas, 11-ago-2026). Usar analitica.v_oficial_componente / v_oficial_municipio / v_oficial_region para 2026 (solo 3 componentes: Captación, Dinamismo, Tecnificación). Sin fuente para sexo/edad en 2023-2025 a nivel municipio.")
        ws.cell(row=row, column=5, value="Fuera de la BD (2023-25) / Parcial (2026)")
        ws.cell(row=row, column=6, value="No asumir ni reconstruir. v_oficial_componente solo trae H/M a nivel ESTATAL, no por municipio ni región. Si se necesita H/M por municipio, hay que pedir que se agregue al ETL — no existe en ningún lugar de la BD hoy.")
    if "entrega parcial por concepto" in rubro:
        ws.cell(row=row, column=4, value="analitica.v_oficial_componente / v_oficial_municipio / v_oficial_region — VERIFICADO: cubre únicamente 3 componentes 2026 (Captación y Almacenamiento de Agua, Dinamismo Agroalimentario, Tecnificación del Riego), no los ~20 programas que sí tiene resumen_estatal para años anteriores.")
        ws.cell(row=row, column=5, value="Parcial (solo 3 de ~20 programas)")
        ws.cell(row=row, column=6, value="El resto de programas 2026 (Municipalizado, Maíz Blanco, Bordería, etc.) no está cargado todavía en ninguna tabla/vista. No inventar cifras para ellos.")
style_rows(ws, 6)

# ---- New sheet: Cobertura por municipio (2023-2025, apoyo_municipio) ----
if "Cobertura 18 municipios" in wb.sheetnames:
    del wb["Cobertura 18 municipios"]
wsc = wb.create_sheet("Cobertura 18 municipios")
wsc.append(["Municipio","Región","Años con datos (apoyo_municipio)","# Programas distintos","Total apoyos (2023-25)","Inversión total $ (2023-25)","Componentes 2026 en v_oficial_municipio"])
style_header(wsc, 7, height=34)

cov = defaultdict(lambda: {"anios": set(), "programas": set(), "apoyos":0, "total":0.0, "region":""})
with open(BASE+"05_apoyo_municipio_full.csv") as f:
    for row in csv.DictReader(f):
        m = row['municipio_nombre']
        cov[m]["anios"].add(row['anio']); cov[m]["programas"].add(row['programa_nombre'])
        cov[m]["apoyos"] += int(row['numero_apoyos'] or 0); cov[m]["total"] += float(row['total'] or 0)
        cov[m]["region"] = row['region_nombre']

mo = defaultdict(set)
with open(BASE+"14_v_oficial_municipio.csv") as f:
    for row in csv.DictReader(f):
        if row['municipio_proyecto']:
            mo[row['municipio_proyecto']].add(row['componente'])

with open(BASE+"01_municipio.csv") as f:
    all_m = [(row['nombre'], row['region_id']) for row in csv.DictReader(f) if row['region_id']]
region_names = {"10":"SAN JUAN DEL RÍO","11":"CADEREYTA","12":"JALPAN","13":"QUERÉTARO"}

for m, rid in sorted(all_m):
    c = cov.get(m)
    comps = mo.get(m, set())
    wsc.append([
        m, region_names.get(rid,""),
        ",".join(sorted(c["anios"])) if c else "SIN DATOS",
        len(c["programas"]) if c else 0,
        c["apoyos"] if c else 0,
        round(c["total"],0) if c else 0,
        ", ".join(sorted(comps)) if comps else "SIN DATOS 2026",
    ])
style_rows(wsc, 7)
for row in range(2, wsc.max_row+1):
    fill = ok_fill if wsc.cell(row=row, column=3).value != "SIN DATOS" else missing_fill
    wsc.cell(row=row, column=3).fill = fill
    fill2 = ok_fill if "SIN DATOS" not in str(wsc.cell(row=row, column=7).value) else missing_fill
    wsc.cell(row=row, column=7).fill = fill2
widths = [24,16,22,14,16,20,42]
for i,w in enumerate(widths, start=1):
    wsc.column_dimensions[get_column_letter(i)].width = w

note_row = wsc.max_row + 2
wsc.cell(row=note_row, column=1, value="Fuente: CSV exportados por el usuario el 11-ago-2026 desde analitica_export (Docker sedea_db). apoyo_municipio NO tiene filas para 2021, 2022 ni 2026 a nivel municipio (confirmado, no es omisión de este mapa). v_oficial_municipio solo cubre 2026 y solo 3 componentes.").font = Font(name=FONT, italic=True, size=9)

# ---- New sheet: Estatal por año (v_inversion_anual, resumen_estatal) ----
if "Estatal por año" in wb.sheetnames:
    del wb["Estatal por año"]
wse = wb.create_sheet("Estatal por año")
wse.append(["Año","Federal","Estatal","Municipal","Productores","Total"])
style_header(wse, 6)
with open(BASE+"12_v_inversion_anual.csv") as f:
    for row in csv.DictReader(f):
        wse.append([row['anio'], round(float(row['federal']),0), round(float(row['estatal']),0), round(float(row['municipal']),0), round(float(row['productores']),0), round(float(row['total']),0)])
style_rows(wse, 6)
widths_e = [10,18,18,18,18,18]
for i,w in enumerate(widths_e, start=1):
    wse.column_dimensions[get_column_letter(i)].width = w
nr = wse.max_row+2
wse.cell(row=nr, column=1, value="Fuente: analitica.v_inversion_anual. NOTA: no incluye 2026 (la vista no trae ese año todavía); usar v_oficial_componente para 2026 estatal, con la salvedad de que solo cubre 3 de los programas.").font = Font(name=FONT, italic=True, size=9)

wb.save("/sessions/vibrant-intelligent-goldberg/mnt/outputs/work/Mapa_disponibilidad_datos_fichas.xlsx")
print("saved v2")
