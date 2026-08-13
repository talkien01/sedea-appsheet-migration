"""
Motor generador de datos de ficha por municipio.
Combina analitica_export (CSV exportados de Postgres) + Datos_referencia_manual_municipios.xlsx
No inventa datos: todo lo que no tiene fuente queda marcado como PENDIENTE en 'warnings'.
"""
import csv, json, sys
from collections import defaultdict
import openpyxl

BASE = "/sessions/vibrant-intelligent-goldberg/mnt/analitica_export/"
MANUAL = "/sessions/vibrant-intelligent-goldberg/mnt/outputs/work/Datos_referencia_manual_municipios.xlsx"

def cargar_apoyo_municipio():
    rows = []
    with open(BASE+"05_apoyo_municipio_full.csv") as f:
        for r in csv.DictReader(f):
            rows.append(r)
    return rows

def cargar_v_oficial_municipio():
    rows = []
    with open(BASE+"14_v_oficial_municipio.csv") as f:
        for r in csv.DictReader(f):
            rows.append(r)
    return rows

def cargar_manual(hoja, municipio):
    wb = openpyxl.load_workbook(MANUAL, data_only=True)
    ws = wb[hoja]
    headers = [c.value for c in ws[1]]
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = dict(zip(headers, row))
        nombre_col = headers[0]
        if d.get(nombre_col) and municipio in str(d.get(nombre_col)):
            out.append(d)
    return out

def generar(municipio):
    warnings = []
    apoyo = [r for r in cargar_apoyo_municipio() if r['municipio_nombre'] == municipio]
    oficial = [r for r in cargar_v_oficial_municipio() if r['municipio_proyecto'] == municipio]

    # --- 1. Resumen histórico 2023-2025 por programa ---
    historico = sorted(apoyo, key=lambda r: (r['anio'], r['programa_nombre']))
    programas_historicos = set(r['programa_nombre'] for r in apoyo)

    total_apoyos = sum(int(r['numero_apoyos'] or 0) for r in apoyo)
    total_inversion = sum(float(r['total'] or 0) for r in apoyo)
    anios_presentes = sorted(set(r['anio'] for r in apoyo))
    for anio_falt in ["2021", "2022", "2026"]:
        if anio_falt not in anios_presentes:
            warnings.append(f"apoyo_municipio no tiene filas de {anio_falt} para {municipio} (no es hueco solo de este municipio: confirmado que ningún municipio tiene {anio_falt} en esta tabla).")

    # --- 2. Avance 2026 (v_oficial) + warning de programas faltantes ---
    componentes_2026 = set(r['componente'] for r in oficial)
    total_apoyos_2026 = sum(int(r['apoyos'] or 0) for r in oficial)
    total_2026 = sum(float(r['total_dictaminado'] or 0) for r in oficial)

    # Mapear qué programas históricos NO tienen equivalente cargado en 2026
    equivalencias = {
        "DINAMISMO AGROALIMENTARIO": "DINAMISMO AGROALIMENTARIO",
        "TECNIFICACIÓN": "TECNIFICACIÓN DEL RIEGO",
        "CAPTACIÓN Y ALMACENAMIENTO DE AGUA": "CAPTACIÓN Y ALMACENAMIENTO DE AGUA",
    }
    programas_sin_2026 = []
    for p in sorted(programas_historicos):
        comp_esperado = equivalencias.get(p, p)
        if comp_esperado not in componentes_2026:
            programas_sin_2026.append(p)
    if programas_sin_2026:
        warnings.append(
            f"2026 incompleto para {municipio}: faltan cargar/entregar estos programas que sí existen en años anteriores: "
            + ", ".join(programas_sin_2026)
            + ". Acción: pedir al equipo de datos que cargue estos programas 2026 en analitica (tabla apoyo_municipio o una vista equivalente a v_oficial para estos conceptos)."
        )

    # --- 3. Datos manuales (territorio, productos, precipitación, demografía) ---
    territorio = cargar_manual("Territorio", municipio)
    productos = cargar_manual("Productos_top", municipio)
    precipitacion = cargar_manual("Precipitacion", municipio)
    demografia = cargar_manual("Demografia_municipal", municipio)

    def campos_vacios(rows, skip_cols=("Municipio","Fuente / fecha de corte","Grupo de edad","Rank","Año","Producto")):
        vacios = 0
        total = 0
        for row in rows:
            for k, v in row.items():
                if k in skip_cols:
                    continue
                total += 1
                if v in (None, "", " "):
                    vacios += 1
        return vacios, total

    for nombre, rows in [("Territorio", territorio), ("Productos_top", productos), ("Precipitacion", precipitacion), ("Demografia_municipal", demografia)]:
        if not rows:
            warnings.append(f"'{nombre}' no tiene fila para {municipio} en la plantilla manual (Datos_referencia_manual_municipios.xlsx).")
            continue
        vacios, total = campos_vacios(rows)
        if vacios == total and total > 0:
            warnings.append(f"'{nombre}' está sin llenar para {municipio} (fila existe pero todas las celdas de dato están vacías). Fuente sugerida: INEGI/SIAP (territorio y productos), CONAGUA (precipitación), INEGI/SIAP (demografía, según definición del equipo).")
        elif vacios > 0:
            warnings.append(f"'{nombre}' está parcialmente lleno para {municipio}: {vacios} de {total} celdas de dato aún vacías.")

    return {
        "municipio": municipio,
        "historico_2023_2025": historico,
        "total_apoyos_historico": total_apoyos,
        "total_inversion_historico": round(total_inversion, 2),
        "anios_presentes": anios_presentes,
        "avance_2026": oficial,
        "total_apoyos_2026": total_apoyos_2026,
        "total_2026": round(total_2026, 2),
        "programas_sin_2026": programas_sin_2026,
        "territorio": territorio,
        "productos": productos,
        "precipitacion": precipitacion,
        "demografia": demografia,
        "warnings": warnings,
    }

if __name__ == "__main__":
    m = sys.argv[1] if len(sys.argv) > 1 else "PEDRO ESCOBEDO"
    data = generar(m)
    print(json.dumps(data, ensure_ascii=False, indent=2))
