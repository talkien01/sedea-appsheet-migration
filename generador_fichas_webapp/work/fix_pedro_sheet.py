import csv, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE = "/sessions/vibrant-intelligent-goldberg/mnt/analitica_export/"
FONT="Arial"
header_fill = PatternFill("solid", fgColor="366092")
header_font = Font(name=FONT, bold=True, color="FFFFFF", size=11)
normal_font = Font(name=FONT, size=10)
thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = openpyxl.load_workbook("/sessions/vibrant-intelligent-goldberg/mnt/outputs/work/Mapa_disponibilidad_datos_fichas.xlsx")
del wb["Pedro Escobedo - datos crudos"]
ws = wb.create_sheet("Pedro Escobedo - datos crudos", 1)
ws.append(["Año","Programa","Apoyos","Apoyo Federal","Apoyo Estatal","Apoyo Municipal","Aportación Productores","Total","Fuente (hoja original)"])
for c in range(1,10):
    cell = ws.cell(row=1, column=c); cell.font=header_font; cell.fill=header_fill; cell.alignment=Alignment(wrap_text=True, vertical="center")
ws.row_dimensions[1].height = 30
ws.freeze_panes="A2"

with open(BASE+"05_apoyo_municipio_full.csv") as f:
    rows = [r for r in csv.DictReader(f) if r['municipio_nombre']=="PEDRO ESCOBEDO"]
rows.sort(key=lambda r: (r['anio'], r['programa_nombre']))
for r in rows:
    ws.append([r['anio'], r['programa_nombre'], int(r['numero_apoyos']), float(r['apoyo_federal']), float(r['apoyo_estatal']), float(r['apoyo_municipal']), round(float(r['aportacion_productor']),2), round(float(r['total']),2), r['fuente_hoja']])

# 2026 v_oficial rows for Pedro Escobedo
with open(BASE+"14_v_oficial_municipio.csv") as f:
    orows = [r for r in csv.DictReader(f) if r['municipio_proyecto']=="PEDRO ESCOBEDO"]
for r in orows:
    ws.append([r['anio'], r['componente']+" (v_oficial: solicitudes="+r['solicitudes']+", apoyos="+r['apoyos']+")", "", "", float(r['estatal_dictaminado'] or 0), "", "", float(r['total_dictaminado'] or 0), "v_oficial_municipio"])

for row in range(2, ws.max_row+1):
    for col in range(1,10):
        cell = ws.cell(row=row, column=col)
        cell.font = normal_font
        cell.border = border
        cell.alignment = Alignment(wrap_text=True, vertical="top")
widths=[8,26,10,14,14,14,16,14,26]
for i,w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width=w

nr = ws.max_row+2
ws.cell(row=nr, column=1, value="Fuente: analitica.apoyo_municipio (filas 2023-2025) + analitica.v_oficial_municipio (fila 2026, solo 3 componentes). Exportado por el usuario el 11-ago-2026 desde el Docker sedea_db. Estas cifras SUSTITUYEN a las extraídas manualmente del Drive en la primera versión de este archivo — la base de datos es ahora la fuente autorizada.").font = Font(name=FONT, italic=True, size=9)
ws.cell(row=nr+1, column=1, value="NO existen filas para Pedro Escobedo en 2021, 2022 (a nivel municipio) — confirmado, no es un hueco de este reporte sino de la base misma.").font = Font(name=FONT, italic=True, size=9)

wb.save("/sessions/vibrant-intelligent-goldberg/mnt/outputs/work/Mapa_disponibilidad_datos_fichas.xlsx")
print("done", ws.max_row)
