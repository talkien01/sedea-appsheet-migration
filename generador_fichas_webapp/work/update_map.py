import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.load_workbook("/sessions/vibrant-intelligent-goldberg/mnt/outputs/work/Mapa_disponibilidad_datos_fichas.xlsx")
ws = wb["Mapa de disponibilidad"]

FONT = "Arial"
normal_font = Font(name=FONT, size=10)
ok_fill = PatternFill("solid", fgColor="C6EFCE")
partial_fill = PatternFill("solid", fgColor="FFEB9C")
missing_fill = PatternFill("solid", fgColor="FFC7CE")
thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# Map: row text fragment (col A) -> (nueva fuente col D, nuevo estado col E, nueva nota col F)
updates = {
 "Resumen de apoyos por tipo (Productivo/Emergente) y aportación F/E/M/Productores|Estatal":
   ("Postgres: analitica.resumen_estatal JOIN analitica.programa (o vista v_ficha_programa_anio)", "Disponible en BD", "Ya no depende del Drive: se puede generar con SQL directo. Falta homologar 'categoria' de programa a Productivo/Emergente."),
 "Resumen de apoyos por tipo (Productivo/Emergente) y aportación F/E/M/Productores|Regional SJR":
   ("Postgres: analitica.apoyo_municipio JOIN municipio JOIN region (filtro region_id) o vista v_ficha_municipio agregada por región", "Disponible en BD", "Agregar apoyo_municipio por region_id en vez de listar municipios a mano."),
 "Resumen de apoyos por tipo (Productivo/Emergente) y aportación F/E/M/Productores|Amealco":
   ("Postgres: analitica.apoyo_municipio WHERE municipio_id = (Amealco de Bonfil)", "Disponible en BD", "Coincide con hoja 'Amealco' del Drive; la BD es ahora la fuente primaria."),
 "Resumen de apoyos por tipo (Productivo/Emergente) y aportación F/E/M/Productores|Pedro Escobedo":
   ("Postgres: analitica.apoyo_municipio WHERE municipio_id = (Pedro Escobedo)", "Disponible en BD", "Misma consulta que Amealco, solo cambia el filtro de municipio. Ya no es necesario reconstruir a mano desde el Drive."),
 "Resumen de apoyos por tipo (Productivo/Emergente) y aportación F/E/M/Productores|Otros municipios / regiones (16 municipios, 6 regionales restantes)":
   ("Postgres: analitica.apoyo_municipio (todos los municipio_id) + tabla municipio para catálogo completo", "Disponible en BD", "La BD ya cubre los 18 municipios / 8 regiones si los datos están cargados; falta confirmar que el cargado (ETL) esté completo para todos, no solo SJR/Amealco."),

 "Inversión total 2021-2026 por año (apoyos, beneficiarios, aportaciones, total)|Estatal":
   ("Postgres: analitica.v_inversion_anual", "Disponible en BD", "Vista ya lista, una fila por año."),
 "Inversión total 2021-2026 por año (apoyos, beneficiarios, aportaciones, total)|Regional SJR":
   ("Postgres: analitica.apoyo_municipio agregado por region_id y anio", "Disponible en BD", ""),
 "Inversión total 2021-2026 por año (apoyos, beneficiarios, aportaciones, total)|Amealco":
   ("Postgres: analitica.apoyo_municipio agregado por municipio_id y anio", "Disponible en BD", ""),
 "Inversión total 2021-2026 por año (apoyos, beneficiarios, aportaciones, total)|Pedro Escobedo":
   ("Postgres: analitica.apoyo_municipio agregado por municipio_id y anio (mismo patrón que Amealco)", "Disponible en BD", "Se resuelve solo: ya no falta 'construir una hoja equivalente a Amealco', es la misma consulta con otro municipio_id."),

 "Extensión territorial y superficie agrícola (riego/temporal)|Estatal":
   ("NO está en el esquema Postgres (schema analitica no tiene tabla de territorio/superficie)", "Fuera de la BD", "Confirmado con el esquema completo (116/116 columnas revisadas): no existe tabla de territorio. Sigue siendo dato manual (INEGI/SIAP) a mantener aparte."),
 "Extensión territorial y superficie agrícola (riego/temporal)|Amealco":
   ("NO está en Postgres", "Fuera de la BD", "Igual que estatal: hay que seguir manteniendo esto fuera de la base, como tabla de referencia estática por municipio."),
 "Extensión territorial y superficie agrícola (riego/temporal)|Pedro Escobedo y resto de municipios":
   ("NO está en Postgres", "Fuera de la BD", "Se recomienda crear una tabla chica aparte (territorio_municipio) con estos datos si se quiere automatizar también; mientras tanto, tabla de referencia manual."),

 "Top 5-6 productos (superficie, volumen, valor) por nivel|Estatal":
   ("NO está en Postgres", "Fuera de la BD", "Confirmado: no hay tabla de producción agrícola/pecuaria por producto en el esquema. Fuente SIAP externa."),
 "Top 5-6 productos (superficie, volumen, valor) por nivel|Amealco":
   ("NO está en Postgres", "Fuera de la BD", ""),
 "Top 5-6 productos (superficie, volumen, valor) por nivel|Pedro Escobedo y resto":
   ("NO está en Postgres", "Fuera de la BD", ""),

 "Precipitación mensual por año (histórico ~2012-2026)|Estatal":
   ("NO está en Postgres", "Fuera de la BD", "Confirmado: no hay tabla de precipitación. Sigue viviendo en los xlsx de CONAGUA del Drive."),
 "Precipitación mensual por año (histórico ~2012-2026)|Amealco / Pedro Escobedo / municipios":
   ("NO está en Postgres", "Fuera de la BD", ""),

 "Distribución de beneficiarios por sexo y grupo de edad|Estatal":
   ("Postgres: analitica.beneficiarios_demografia (por programa/año) y/o analitica.v_oficial_componente (benef_hombres, benef_mujeres)", "Disponible en BD", ""),
 "Distribución de beneficiarios por sexo y grupo de edad|Regional SJR":
   ("Postgres: analitica.v_oficial_region (benef_hombres/mujeres no están en esta vista, pero sí en v_oficial_componente cruzado con region si se ajusta la vista)", "Parcial en BD", "v_oficial_region trae solicitudes/apoyos/montos por región pero no columnas H/M; v_oficial_componente sí trae H/M pero a nivel estatal, no por región. Falta una vista que cruce ambas dimensiones."),
 "Distribución de beneficiarios por sexo y grupo de edad|Amealco":
   ("Igual que Regional: falta vista municipio x género", "Parcial en BD", "v_oficial_municipio no trae columnas de género; habría que pedir se agregue o derivarlo de accion/beneficiario si existe esa liga."),
 "Distribución de beneficiarios por sexo y grupo de edad|Pedro Escobedo":
   ("Igual que Amealco", "Parcial en BD", "beneficiarios_demografia NO tiene columna municipio_id (confirmado por FK): el desglose por sexo/edad hoy solo existe a nivel programa/año, no por municipio. Es una limitación real del modelo, no solo falta de carga."),

 "Situación de almacenamientos / presas|Estatal":
   ("NO está en Postgres", "Fuera de la BD", "Sigue siendo el PDF/xlsx de CONAGUA/presas del Drive."),

 "Programa institucional 'Apoyo al Campo Queretano' 2026 — entrega parcial por concepto|Estatal":
   ("Postgres: analitica.v_oficial_componente (columna 'apoyos_positivos', filtrable por año 2026)", "Disponible en BD", "Esta vista parece ser justo la fuente de la sección 'entrega parcial' de la ficha (solicitudes vs apoyos vs dictaminado)."),
 "Programa institucional 'Apoyo al Campo Queretano' 2026 — entrega parcial por concepto|Regional SJR / Amealco":
   ("Postgres: analitica.v_oficial_region y analitica.v_oficial_municipio", "Disponible en BD", ""),
 "Programa institucional 'Apoyo al Campo Queretano' 2026 — entrega parcial por concepto|Pedro Escobedo y resto":
   ("Postgres: analitica.v_oficial_municipio (mismo patrón, cualquier municipio)", "Disponible en BD", "Ya no depende de encontrar el archivo por municipio en Drive; es una fila de esta vista."),

 "Metodología / definición de rubros|Todos":
   ("Esquema Postgres schema 'analitica' — 10 tablas base + 6 vistas, confirmado completo vía information_schema (116/116 columnas) y FKs verificadas", "Documentado", "Ver hoja 'Esquema Postgres' de este mismo archivo para el detalle tabla por tabla."),
}

changed = 0
for row in range(2, ws.max_row+1):
    rubro = ws.cell(row=row, column=1).value
    nivel = ws.cell(row=row, column=2).value
    key = f"{rubro}|{nivel}"
    if key in updates:
        fuente, estado, nota = updates[key]
        ws.cell(row=row, column=4, value=fuente)
        ws.cell(row=row, column=5, value=estado)
        ws.cell(row=row, column=6, value=nota)
        changed += 1

print("filas actualizadas:", changed)

# re-style status column colors
for row in range(2, ws.max_row+1):
    for col in range(1,7):
        c = ws.cell(row=row, column=col)
        c.font = normal_font
        c.alignment = Alignment(wrap_text=True, vertical="top")
        c.border = border
    status_val = str(ws.cell(row=row, column=5).value or "")
    fill = None
    if "Disponible" in status_val:
        fill = ok_fill
    elif "Parcial" in status_val or "confirmar" in status_val or "Documentado" in status_val:
        fill = partial_fill
    elif "Fuera de la BD" in status_val or "No localizado" in status_val or "Pendiente" in status_val:
        fill = missing_fill
    if fill:
        ws.cell(row=row, column=5).fill = fill

# ---- New sheet: esquema Postgres ----
if "Esquema Postgres" in wb.sheetnames:
    del wb["Esquema Postgres"]
ws3 = wb.create_sheet("Esquema Postgres")
header_fill = PatternFill("solid", fgColor="366092")
header_font = Font(name=FONT, bold=True, color="FFFFFF", size=11)
ws3.append(["Tabla/Vista", "Tipo", "Columna", "Tipo de dato", "Llave foránea"])
for c in range(1,6):
    cell = ws3.cell(row=1, column=c); cell.font=header_font; cell.fill=header_fill; cell.alignment=Alignment(vertical="center")

schema = [
 ("region","Tabla catálogo",[("region_id","smallint",""),("nombre","text","")]),
 ("municipio","Tabla catálogo",[("municipio_id","smallint",""),("nombre","text",""),("region_id","smallint","region.region_id")]),
 ("municipio_alias","Tabla catálogo",[("alias","text",""),("municipio_id","smallint","municipio.municipio_id")]),
 ("programa","Tabla catálogo",[("programa_id","integer",""),("nombre","text",""),("categoria","text","")]),
 ("programa_alias","Tabla catálogo",[("alias","text",""),("programa_id","integer","programa.programa_id")]),
 ("apoyo_municipio","Tabla hechos",[("apoyo_id","bigint",""),("anio","smallint",""),("programa_id","integer","programa.programa_id"),("municipio_id","smallint","municipio.municipio_id"),("numero_apoyos","integer",""),("apoyo_federal","numeric",""),("apoyo_estatal","numeric",""),("apoyo_municipal","numeric",""),("aportacion_productor","numeric",""),("total","numeric",""),("fuente_archivo","text",""),("fuente_hoja","text",""),("cargado_en","timestamptz","")]),
 ("apoyo_metrica","Tabla hechos",[("apoyo_id","bigint","apoyo_municipio.apoyo_id"),("metrica","text",""),("valor","numeric","")]),
 ("resumen_estatal","Tabla hechos",[("resumen_id","bigint",""),("anio","smallint",""),("programa_id","integer","programa.programa_id"),("beneficiarios","integer",""),("beneficiarios_indir","integer",""),("numero_apoyos","integer",""),("apoyo_federal","numeric",""),("apoyo_estatal","numeric",""),("apoyo_municipal","numeric",""),("aportacion_productor","numeric",""),("total","numeric",""),("fuente_archivo","text",""),("cargado_en","timestamptz","")]),
 ("accion","Tabla hechos (detalle de obra)",[("accion_id","bigint",""),("anio","smallint",""),("programa_id","integer","programa.programa_id"),("subprograma","text",""),("responsable","text",""),("num_obra","text",""),("fecha","date",""),("obra_accion","text",""),("unidad","text",""),("cantidad","numeric",""),("municipio_id","smallint","municipio.municipio_id"),("localidad","text",""),("inv_federal","numeric",""),("inv_estatal","numeric",""),("inv_beneficiario","numeric",""),("inv_total","numeric",""),("num_beneficiarios","integer",""),("clasificacion","text",""),("metricas","jsonb",""),("beneficiario_id","bigint",""),("fuente_archivo","text",""),("cargado_en","timestamptz","")]),
 ("beneficiarios_demografia","Tabla hechos (SIN municipio)",[("id","bigint",""),("anio","smallint",""),("programa_id","integer","programa.programa_id"),("genero","text",""),("rango_edad","text",""),("cantidad","integer","")]),
 ("v_ficha_municipio","Vista",[("anio","smallint",""),("municipio","text",""),("programa","text",""),("numero_apoyos","integer",""),("apoyo_estatal","numeric",""),("apoyo_municipal","numeric",""),("aportacion_productor","numeric",""),("total","numeric","")]),
 ("v_ficha_programa_anio","Vista",[("anio","smallint",""),("programa","text",""),("beneficiarios","integer",""),("numero_apoyos","integer",""),("apoyo_federal","numeric",""),("apoyo_estatal","numeric",""),("apoyo_municipal","numeric",""),("aportacion_productor","numeric",""),("total","numeric","")]),
 ("v_inversion_anual","Vista",[("anio","smallint",""),("federal","numeric",""),("estatal","numeric",""),("municipal","numeric",""),("productores","numeric",""),("total","numeric","")]),
 ("v_oficial_componente","Vista (con género)",[("anio","smallint",""),("componente","text",""),("solicitudes","bigint",""),("apoyos","bigint",""),("apoyos_positivos","bigint",""),("benef_hombres","bigint",""),("benef_mujeres","bigint",""),("estatal_solicitado","numeric",""),("estatal_dictaminado","numeric",""),("total_dictaminado","numeric","")]),
 ("v_oficial_municipio","Vista",[("anio","smallint",""),("municipio_proyecto","text",""),("componente","text",""),("solicitudes","bigint",""),("apoyos","bigint",""),("estatal_dictaminado","numeric",""),("total_dictaminado","numeric","")]),
 ("v_oficial_region","Vista",[("anio","smallint",""),("regional","text",""),("componente","text",""),("solicitudes","bigint",""),("apoyos","bigint",""),("estatal_dictaminado","numeric",""),("total_dictaminado","numeric","")]),
]
for tabla, tipo, cols in schema:
    for colname, dtype, fk in cols:
        ws3.append([tabla, tipo, colname, dtype, fk])
for row in range(2, ws3.max_row+1):
    for col in range(1,6):
        c = ws3.cell(row=row, column=col)
        c.font = normal_font
        c.border = border
w3 = [24,26,22,16,26]
for i,w in enumerate(w3, start=1):
    ws3.column_dimensions[get_column_letter(i)].width = w
ws3.freeze_panes = "A2"

note = ws3.max_row + 2
ws3.cell(row=note, column=1, value="Esquema confirmado completo: 116/116 columnas vistas vía information_schema.columns + FKs verificadas vía information_schema.table_constraints (ambas consultas corridas por el usuario, 11-ago-2026). NO existen tablas de territorio, producción por cultivo/especie, ni precipitación en este esquema — esos 3 bloques de la ficha deben seguir viniendo de fuente externa (Drive / INEGI / SIAP / CONAGUA).").font = Font(name=FONT, italic=True, size=9)

wb.save("/sessions/vibrant-intelligent-goldberg/mnt/outputs/work/Mapa_disponibilidad_datos_fichas.xlsx")
print("saved")
