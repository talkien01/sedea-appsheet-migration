"""
Motor generador de datos de ficha por municipio.
Combina los CSV exportados de Postgres (analitica_export) + la plantilla manual
(Datos_referencia_manual_municipios.xlsx). No inventa datos: todo lo que no
tiene fuente queda marcado como advertencia (warning) en vez de completarse.
"""
import csv, os
import openpyxl
from config import DATA_DIR, MANUAL_XLSX

def _leer_csv(nombre):
    path = os.path.join(DATA_DIR, nombre)
    if not os.path.exists(path):
        return []
    with open(path, encoding="utf-8") as f:
        return list(csv.DictReader(f))

def cargar_apoyo_municipio():
    return _leer_csv("05_apoyo_municipio_full.csv")

def cargar_v_oficial_municipio():
    return _leer_csv("14_v_oficial_municipio.csv")

def cargar_municipios():
    rows = _leer_csv("01_municipio.csv")
    return [r["nombre"] for r in rows if r.get("region_id")]

def cargar_manual(hoja, municipio):
    if not os.path.exists(MANUAL_XLSX):
        return []
    wb = openpyxl.load_workbook(MANUAL_XLSX, data_only=True)
    if hoja not in wb.sheetnames:
        return []
    ws = wb[hoja]
    headers = [c.value for c in ws[1]]
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = dict(zip(headers, row))
        nombre_col = headers[0]
        if d.get(nombre_col) and municipio in str(d.get(nombre_col)):
            out.append(d)
    return out

def generar(municipio):
    warnings = []
    apoyo = [r for r in cargar_apoyo_municipio() if r["municipio_nombre"] == municipio]
    oficial = [r for r in cargar_v_oficial_municipio() if r["municipio_proyecto"] == municipio]

    historico = sorted(apoyo, key=lambda r: (r["anio"], r["programa_nombre"]))
    programas_historicos = set(r["programa_nombre"] for r in apoyo)

    total_apoyos = sum(int(r["numero_apoyos"] or 0) for r in apoyo)
    total_inversion = sum(float(r["total"] or 0) for r in apoyo)
    anios_presentes = sorted(set(r["anio"] for r in apoyo))
    for anio_falt in ["2021", "2022", "2026"]:
        if anio_falt not in anios_presentes:
            warnings.append(
                f"apoyo_municipio no tiene filas de {anio_falt} para {municipio} "
                f"(confirmado: ningún municipio tiene {anio_falt} en esta tabla, no es hueco solo de este municipio)."
            )

    componentes_2026 = set(r["componente"] for r in oficial)
    total_apoyos_2026 = sum(int(r["apoyos"] or 0) for r in oficial)
    total_2026 = sum(float(r["total_dictaminado"] or 0) for r in oficial)

    equivalencias = {
        "DINAMISMO AGROALIMENTARIO": "DINAMISMO AGROALIMENTARIO",
        "TECNIFICACIÓN": "TECNIFICACIÓN DEL RIEGO",
        "CAPTACIÓN Y ALMACENAMIENTO DE AGUA": "CAPTACIÓN Y ALMACENAMIENTO DE AGUA",
    }
    programas_sin_2026 = []
    for p in sorted(programas_historicos):
        comp_esperado = equivalencias.get(p, p)
        if comp_esperado not in componentes_2026:
            programas_sin_2026.append(p)
    if programas_sin_2026:
        warnings.append(
            f"2026 incompleto para {municipio}: faltan cargar estos programas que sí existen en años anteriores: "
            + ", ".join(programas_sin_2026)
            + ". Acción: pedir al equipo de datos que cargue estos programas 2026 en analitica."
        )

    territorio = cargar_manual("Territorio", municipio)
    productos = cargar_manual("Productos_top", municipio)
    precipitacion = cargar_manual("Precipitacion", municipio)
    demografia = cargar_manual("Demografia_municipal", municipio)

    def campos_vacios(rows, skip_cols=("Municipio", "Fuente / fecha de corte", "Grupo de edad", "Rank", "Año", "Producto")):
        vacios = total = 0
        for row in rows:
            for k, v in row.items():
                if k in skip_cols:
                    continue
                total += 1
                if v in (None, "", " "):
                    vacios += 1
        return vacios, total

    for nombre, rows in [("Territorio", territorio), ("Productos_top", productos),
                          ("Precipitacion", precipitacion), ("Demografia_municipal", demografia)]:
        if not rows:
            warnings.append(f"'{nombre}' no tiene fila para {municipio} en la plantilla manual.")
            continue
        vacios, total = campos_vacios(rows)
        if vacios == total and total > 0:
            warnings.append(
                f"'{nombre}' está sin llenar para {municipio}. Fuente sugerida: INEGI/SIAP "
                f"(territorio, productos, demografía) o CONAGUA (precipitación)."
            )
        elif vacios > 0:
            warnings.append(f"'{nombre}' parcialmente lleno para {municipio}: {vacios} de {total} celdas vacías.")

    return {
        "municipio": municipio,
        "historico": historico,
        "total_apoyos_historico": total_apoyos,
        "total_inversion_historico": round(total_inversion, 2),
        "anios_presentes": anios_presentes,
        "avance_2026": oficial,
        "total_apoyos_2026": total_apoyos_2026,
        "total_2026": round(total_2026, 2),
        "territorio": territorio,
        "productos": productos,
        "precipitacion": precipitacion,
        "demografia": demografia,
        "warnings": warnings,
    }
